"""Router T&C — Clientes corporativos. Prefijo: /tc"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Optional

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlmodel import Session, col, select

from app.core.deps import require_permission
from app.database import get_db
from app.models.sede import Sede
from app.models.user import User
from app.personal_database import (
    PtcCliente,
    PtcClienteAsignacion,
    PtcPersona,
    get_personal_engine,
)

router = APIRouter(prefix="/tc", tags=["T&C Clientes"])

require_tc = require_permission("mod_tc")
require_tc_editar = require_permission("mod_tc_editar")

_ACTIVO = "Activo"


class AsignacionBody(BaseModel):
    sede_id: int
    persona_id: Optional[int] = None


class ClienteBody(BaseModel):
    client_no: str = Field(min_length=1, max_length=50)
    dume_no: str = Field(default="", max_length=50)
    nombre: str = Field(min_length=1, max_length=200)
    activo: bool = True
    asignaciones: list[AsignacionBody] = Field(default_factory=list)


class ClienteUpdateBody(BaseModel):
    dume_no: Optional[str] = None
    nombre: Optional[str] = None
    activo: Optional[bool] = None
    asignaciones: Optional[list[AsignacionBody]] = None


def _get_personal_db():
    with Session(get_personal_engine()) as session:
        yield session


def _client_field(row: dict, keys: list[str]) -> str:
    for k in keys:
        for rk, rv in row.items():
            if str(rk).strip().lower() == k.lower():
                return str(rv or "").strip()
    return ""


def _cliente_dict(c: PtcCliente, db: Session, main_db: Session) -> dict:
    asigs = db.exec(
        select(PtcClienteAsignacion).where(PtcClienteAsignacion.cliente_id == c.id)
    ).all()
    asignaciones: dict[str, dict] = {}
    for a in asigs:
        persona = db.get(PtcPersona, a.persona_id) if a.persona_id else None
        sede = main_db.get(Sede, a.sede_id) if a.sede_id else None
        asignaciones[str(a.sede_id)] = {
            "sede_id": a.sede_id,
            "sede_nombre": sede.name if sede else "",
            "persona_id": a.persona_id,
            "persona_nombre": persona.nombre if persona else "",
        }
    return {
        "id": c.id,
        "client_no": c.client_no,
        "dume_no": c.dume_no,
        "nombre": c.nombre,
        "activo": c.activo,
        "asignaciones": asignaciones,
        "created_at": c.created_at.isoformat(),
        "updated_at": c.updated_at.isoformat(),
    }


def _sync_asignaciones(db: Session, cliente_id: int, items: list[AsignacionBody]) -> None:
    existing = db.exec(
        select(PtcClienteAsignacion).where(PtcClienteAsignacion.cliente_id == cliente_id)
    ).all()
    for row in existing:
        db.delete(row)
    for item in items:
        if item.persona_id:
            db.add(PtcClienteAsignacion(
                cliente_id=cliente_id,
                sede_id=item.sede_id,
                persona_id=item.persona_id,
            ))


def _stats(clientes: list[PtcCliente], db: Session) -> dict:
    total = len(clientes)
    asignados = 0
    for c in clientes:
        rows = db.exec(
            select(PtcClienteAsignacion).where(
                PtcClienteAsignacion.cliente_id == c.id,
                col(PtcClienteAsignacion.persona_id).is_not(None),
            )
        ).all()
        if rows:
            asignados += 1
    return {
        "total": total,
        "asignados": asignados,
        "pendientes": total - asignados,
    }


@router.get("/clientes")
def listar_clientes(
    q: Optional[str] = Query(default=None),
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=200, le=500),
    db: Session = Depends(_get_personal_db),
    main_db: Session = Depends(get_db),
    _: User = Depends(require_tc),
):
    query = select(PtcCliente).where(PtcCliente.activo == True)  # noqa: E712
    if q:
        term = f"%{q.lower()}%"
        query = query.where(
            col(PtcCliente.nombre).ilike(term)
            | col(PtcCliente.client_no).ilike(term)
            | col(PtcCliente.dume_no).ilike(term)
        )
    all_rows = db.exec(query.order_by(col(PtcCliente.nombre))).all()
    stats = _stats(all_rows, db)
    page = all_rows[skip : skip + limit]
    return {
        **stats,
        "items": [_cliente_dict(c, db, main_db) for c in page],
    }


@router.post("/clientes", status_code=status.HTTP_201_CREATED)
def crear_cliente(
    body: ClienteBody,
    db: Session = Depends(_get_personal_db),
    main_db: Session = Depends(get_db),
    _: User = Depends(require_tc_editar),
):
    dup = db.exec(
        select(PtcCliente).where(PtcCliente.client_no == body.client_no.strip())
    ).first()
    if dup:
        raise HTTPException(status_code=409, detail="Ya existe un cliente con ese número.")

    c = PtcCliente(
        client_no=body.client_no.strip(),
        dume_no=body.dume_no.strip(),
        nombre=body.nombre.strip(),
        activo=body.activo,
    )
    db.add(c)
    db.flush()
    _sync_asignaciones(db, c.id, body.asignaciones)
    db.commit()
    db.refresh(c)
    return _cliente_dict(c, db, main_db)


@router.put("/clientes/{cliente_id}")
def actualizar_cliente(
    cliente_id: int,
    body: ClienteUpdateBody,
    db: Session = Depends(_get_personal_db),
    main_db: Session = Depends(get_db),
    _: User = Depends(require_tc_editar),
):
    c = db.get(PtcCliente, cliente_id)
    if not c:
        raise HTTPException(status_code=404, detail="Cliente no encontrado.")

    if body.dume_no is not None:
        c.dume_no = body.dume_no.strip()
    if body.nombre is not None:
        c.nombre = body.nombre.strip()
    if body.activo is not None:
        c.activo = body.activo
    if body.asignaciones is not None:
        _sync_asignaciones(db, c.id, body.asignaciones)
    c.updated_at = datetime.utcnow()
    db.add(c)
    db.commit()
    db.refresh(c)
    return _cliente_dict(c, db, main_db)


@router.delete("/clientes/{cliente_id}", status_code=status.HTTP_204_NO_CONTENT)
def eliminar_cliente(
    cliente_id: int,
    db: Session = Depends(_get_personal_db),
    _: User = Depends(require_tc_editar),
):
    c = db.get(PtcCliente, cliente_id)
    if not c:
        raise HTTPException(status_code=404, detail="Cliente no encontrado.")
    c.activo = False
    c.updated_at = datetime.utcnow()
    db.add(c)
    db.commit()


@router.get("/clientes/plantilla")
def descargar_plantilla_clientes(_: User = Depends(require_tc)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"
    ws.append(["No de Cliente", "No de DUME", "Nombre o razón social"])
    ws.append(["CLI-001", "DUME-123", "Ejemplo Cliente S.A.S."])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="Plantilla_Clientes_TYC.xlsx"'},
    )


@router.post("/clientes/import/excel")
async def importar_clientes_excel(
    file: UploadFile = File(...),
    db: Session = Depends(_get_personal_db),
    main_db: Session = Depends(get_db),
    _: User = Depends(require_tc_editar),
):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=422, detail="Formato no soportado. Usa .xlsx")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=422, detail="No se pudo leer el Excel.") from exc

    ws = wb.active
    headers = [str(c.value or "").strip() for c in ws[1]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    created = updated = skipped = 0

    for row in rows:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            skipped += 1
            continue
        row_dict = dict(zip(headers, row))
        client_no = _client_field(
            row_dict,
            ["No de Cliente", "No Cliente", "Cliente", "N° Cliente", "Numero de Cliente"],
        )
        dume_no = _client_field(
            row_dict,
            ["No de DUME", "No DUME", "DUME", "N° DUME", "Numero de DUME"],
        )
        nombre = _client_field(
            row_dict,
            ["Nombre o razón social", "Nombre", "Razon social", "Nombre o razon social"],
        )

        if not client_no or not nombre:
            skipped += 1
            continue

        existing = db.exec(
            select(PtcCliente).where(PtcCliente.client_no == client_no)
        ).first()
        if existing:
            existing.dume_no = dume_no
            existing.nombre = nombre
            existing.activo = True
            existing.updated_at = datetime.utcnow()
            db.add(existing)
            updated += 1
        else:
            db.add(PtcCliente(
                client_no=client_no,
                dume_no=dume_no,
                nombre=nombre,
                activo=True,
            ))
            created += 1

    db.commit()
    return {"created": created, "updated": updated, "skipped": skipped}


@router.get("/clientes/analistas")
def listar_analistas(
    db: Session = Depends(_get_personal_db),
    _: User = Depends(require_tc),
):
    """Personas activas cuyo cargo contiene 'analista' y 'operac' (Analista de Operaciones)."""
    personas = db.exec(
        select(PtcPersona).where(PtcPersona.estado == _ACTIVO)
    ).all()
    from app.personal_database import PtcCargo
    result = []
    for p in personas:
        if not p.cargo_id:
            continue
        cargo = db.get(PtcCargo, p.cargo_id)
        if not cargo:
            continue
        cn = cargo.nombre.lower()
        if "analista" in cn and "operac" in cn:
            result.append({
                "id": p.id,
                "nombre": p.nombre,
                "sede_id": p.sede_id,
            })
    return result
