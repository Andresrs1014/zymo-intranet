import subprocess
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, status
from fastapi.responses import FileResponse
from pydantic import BaseModel
from sqlmodel import Session, func, select

from app.core.deps import get_current_user, require_compras
from app.database import get_db
from app.models.oc import CotizacionProveedor, OrdenCompra, SolicitudOC
from app.models.user import User
from app.oc_database import get_oc_db
from app.services.historial import registrar_cambio_estado

router = APIRouter(tags=["OC - Documentos"])

OC_DOCS_DIR = Path("/app/data/oc_docs")


# ── Schemas ───────────────────────────────────────────────────────────────────

class MarcarEnviadaPayload(BaseModel):
    email_proveedor: Optional[str] = None


class OrdenCompraRead(BaseModel):
    id: uuid.UUID
    solicitud_id: uuid.UUID
    cotizacion_id: uuid.UUID
    numero_oc: str
    pdf_path: Optional[str]
    enviada_proveedor: bool
    enviada_coordinador: bool
    email_proveedor: Optional[str]
    created_at: datetime

    class Config:
        from_attributes = True


# ── Helpers ───────────────────────────────────────────────────────────────────

def _format_cop(value: Optional[float]) -> str:
    if value is None:
        return "N/A"
    return f"${value:,.0f} COP"


_PLATFORMS_DIR = Path(__file__).parent.parent.parent / "platforms"
_STATIC_DIR = Path(__file__).parent.parent.parent / "static"

_SLUG_MAP = {
    "logimat": "logimat",
    "logimat s.a.s.": "logimat",
    "imccargo": "imccargo",
    "imc cargo": "imccargo",
    "imc cargo international": "imccargo",
    "imc cargo international s.a.s.": "imccargo",
    "imcdep": "imcdep",
    "imc deposito": "imcdep",
    "imc depósito": "imcdep",
    "imc deposito s.a.s.": "imcdep",
    "imc depósito s.a.s.": "imcdep",
}


def _load_platform_config(plataforma: Optional[str]) -> dict:
    """Carga la config de la plataforma desde app/platforms/{slug}/config.json."""
    import json
    slug = _SLUG_MAP.get((plataforma or "").lower().strip(), "logimat")
    config_path = _PLATFORMS_DIR / slug / "config.json"
    if config_path.exists():
        return json.loads(config_path.read_text(encoding="utf-8"))
    return {
        "nombre": "LOGIMAT S.A.S.",
        "nit": "830.103.877-6",
        "direccion": "Carrera 106 No. 15A-25 - Manzana 23 LTE 135M",
        "ciudad": "Zona Franca de Bogotá - Colombia",
        "pbx": "(1) 7 44 92 00",
        "email_facturacion": "830103877@factureinbox.co",
        "prefijo_oc": "L",
        "logo": "logimat_logo.png",
    }


def _escribir_plazo_entrega(ws, cfg: dict, plazo_entrega: str) -> None:
    """Detecta el tipo de plazo y escribe en la celda correspondiente del template."""
    import re

    valor = plazo_entrega.strip().upper()

    # INMEDIATA / INMEDIATO
    if "INMEDIATA" in valor or "INMEDIATO" in valor or valor in ("INMEDIATA", "INMEDIATO"):
        if cfg.get("plazo_inmediata_x"):
            ws[cfg["plazo_inmediata_x"]] = "X"
        return

    # Número de días — extraer primer entero del texto
    numeros = re.findall(r"\d+", plazo_entrega)
    if numeros:
        if cfg.get("plazo_dias"):
            ws[cfg["plazo_dias"]] = int(numeros[0])
        return

    # Si llega algo parecido a fecha, escribir en plazo_fecha
    if cfg.get("plazo_fecha"):
        ws[cfg["plazo_fecha"]] = plazo_entrega


def _generar_xlsx(
    numero_oc: str,
    solicitud: SolicitudOC,
    cotizacion: CotizacionProveedor,
    output_path: Path,
    auxiliar_nombre: str = "",
    aprobador_nombre: str = "",
) -> None:
    """Rellena la plantilla Excel de la plataforma con los datos de la OC."""
    import openpyxl
    from zoneinfo import ZoneInfo
    from openpyxl.utils import column_index_from_string
    from openpyxl.worksheet.page import PageMargins

    empresa = _load_platform_config(solicitud.plataforma)
    slug = _SLUG_MAP.get((solicitud.plataforma or "").lower().strip(), "logimat")
    template_path = _PLATFORMS_DIR / slug / empresa.get("template", "template.xlsx")

    wb = openpyxl.load_workbook(str(template_path))
    ws = wb.active

    # ── Logo ──────────────────────────────────────────────────────────────────
    logo_filename = empresa.get("logo")
    if logo_filename:
        logo_path = _PLATFORMS_DIR / slug / logo_filename
        if logo_path.exists():
            from openpyxl.drawing.image import Image as XLImage
            import io
            from PIL import Image as PILImage
            # Convertir a JPEG en memoria — LibreOffice falla al convertir a PDF
            # cuando el XLSX contiene imágenes PNG (especialmente con canal alfa).
            pil_img = PILImage.open(str(logo_path)).convert("RGB")
            buf = io.BytesIO()
            pil_img.save(buf, format="JPEG", quality=95)
            buf.seek(0)
            ws._images = []
            img = XLImage(buf)
            img.width  = empresa.get("logo_width", 150)
            img.height = empresa.get("logo_height", 55)
            img.anchor = empresa.get("logo_anchor", "C3")
            ws.add_image(img)

    fecha_str = datetime.now(ZoneInfo("America/Bogota")).strftime("%d/%m/%Y")
    cfg       = empresa.get("celdas_dinamicas", {})
    items_cfg = empresa.get("items", {})

    # ── Cabecera ──────────────────────────────────────────────────────────────
    if cfg.get("numero_oc"):
        ws[cfg["numero_oc"]] = f"ORDEN DE COMPRA No. {numero_oc}"
    if cfg.get("fecha"):
        ws[cfg["fecha"]] = fecha_str
    if cfg.get("proveedor_nombre"):
        ws[cfg["proveedor_nombre"]] = cotizacion.proveedor_nombre or ""
    if cfg.get("proveedor_nit"):
        ws[cfg["proveedor_nit"]] = cotizacion.proveedor_nit or ""
    if cfg.get("os_ref"):
        ws[cfg["os_ref"]] = f"OS: {solicitud.consecutivo_os}" if solicitud.consecutivo_os else ""
    if cfg.get("cot_ref"):
        ws[cfg["cot_ref"]] = cotizacion.numero_cotizacion_proveedor or ""

    # ── Firmas ────────────────────────────────────────────────────────────────
    if cfg.get("solicita"):
        ws[cfg["solicita"]] = solicitud.solicitante_nombre or ""
    if cfg.get("area_firma"):
        ws[cfg["area_firma"]] = solicitud.area_solicitante or ""
    if cfg.get("elabora"):
        ws[cfg["elabora"]] = auxiliar_nombre
    if cfg.get("aprueba"):
        ws[cfg["aprueba"]] = aprobador_nombre

    # ── Nota / campos operativos ──────────────────────────────────────────────
    if cfg.get("nota"):
        ws[cfg["nota"]] = cotizacion.observaciones or solicitud.observaciones_solicitante or ""

    # Forma de pago
    if cfg.get("forma_pago_x"):
        ws[cfg["forma_pago_x"]] = None
    if cfg.get("forma_pago") and cotizacion.forma_pago:
        ws[cfg["forma_pago"]] = cotizacion.forma_pago
        if cfg.get("forma_pago_x"):
            ws[cfg["forma_pago_x"]] = "X"

    # Plazo
    for _ck in ("plazo_inmediata_x", "plazo_dias", "plazo_fecha"):
        if cfg.get(_ck):
            ws[cfg[_ck]] = None
    if cotizacion.plazo_entrega:
        _escribir_plazo_entrega(ws, cfg, cotizacion.plazo_entrega)

    # Condiciones
    if cfg.get("garantia") and cotizacion.garantia:
        ws[cfg["garantia"]] = cotizacion.garantia
    if cfg.get("anticipo") and cotizacion.anticipo:
        ws[cfg["anticipo"]] = cotizacion.anticipo
    if cfg.get("pago_saldo") and cotizacion.pago_saldo:
        ws[cfg["pago_saldo"]] = cotizacion.pago_saldo

    # ── Ítems ─────────────────────────────────────────────────────────────────
    fila_inicio = items_cfg.get("fila_inicio", 11)
    max_filas   = items_cfg.get("max_filas", 20)
    col_num     = items_cfg.get("col_item_num", "C")
    col_cant    = items_cfg.get("col_cantidad", "D")
    col_ref     = items_cfg.get("col_referencia")
    col_desc    = items_cfg.get("col_descripcion", "F")
    col_vunit   = items_cfg.get("col_valor_unitario", "G")
    col_total   = items_cfg.get("col_total")

    items_a_escribir: list[dict] = cotizacion.items or [{
        "num": 1,
        "descripcion": solicitud.descripcion or "",
        "referencia": solicitud.placa_ficha or "",
        "cantidad": solicitud.cantidad,
        "valor_unitario": cotizacion.valor_unitario or 0,
    }]

    # Construir set de celdas que son parte de rangos fusionados (no top-left)
    # para evitar ValueError al escribir en ellas
    _merged_non_topleft: set[str] = set()
    for merged_range in ws.merged_cells.ranges:
        cells = list(merged_range.cells)
        for r, c in cells[1:]:  # [0] es top-left, el resto no se pueden asignar
            from openpyxl.utils import get_column_letter
            _merged_non_topleft.add(f"{get_column_letter(c)}{r}")

    def _safe_write(addr: str, value) -> None:
        if addr not in _merged_non_topleft:
            ws[addr] = value

    # Limpiar filas del área de ítems
    for fila in range(fila_inicio, fila_inicio + max_filas):
        _safe_write(f"{col_num}{fila}", None)
        _safe_write(f"{col_cant}{fila}", None)
        if col_ref:
            _safe_write(f"{col_ref}{fila}", None)
        _safe_write(f"{col_desc}{fila}", None)
        _safe_write(f"{col_vunit}{fila}", None)
        if col_total:
            _safe_write(f"{col_total}{fila}", None)

    # Escribir ítems (limitado al espacio disponible en el template)
    n_escritos = 0
    for i, item in enumerate(items_a_escribir[:max_filas]):
        fila = fila_inicio + i
        _safe_write(f"{col_num}{fila}", item.get("num") or (i + 1))
        cant = item.get("cantidad")
        _safe_write(f"{col_cant}{fila}", cant if cant is not None else "")
        if col_ref:
            _safe_write(f"{col_ref}{fila}", item.get("referencia") or "")
        _safe_write(f"{col_desc}{fila}", item.get("descripcion") or "")
        vunit = item.get("valor_unitario") or 0
        _safe_write(f"{col_vunit}{fila}", vunit)
        # Fórmula de total por fila — la primera ya la tiene el template,
        # para las extra la escribimos nosotros
        if col_total:
            total_addr = f"{col_total}{fila}"
            if total_addr not in _merged_non_topleft and ws[total_addr].value is None:
                ws[total_addr] = f"=+{col_vunit}{fila}*{col_cant}{fila}"
        n_escritos += 1

    # Si hay más de 1 ítem y el template tiene una celda SUM en col_total,
    # actualizamos su rango para que incluya todas las filas escritas
    if col_total and n_escritos > 1:
        import re as _re
        ultima_fila = fila_inicio + n_escritos - 1
        # Escanear las 15 filas siguientes al área de ítems buscando =SUM(...)
        for scan_row in range(ultima_fila + 1, ultima_fila + 16):
            cell = ws[f"{col_total}{scan_row}"]
            val = cell.value
            if val and isinstance(val, str) and "SUM" in val.upper():
                cell.value = _re.sub(
                    r"SUM\([A-Z]+\d+:[A-Z]+\d+\)",
                    f"SUM({col_total}{fila_inicio}:{col_total}{ultima_fila})",
                    val,
                    flags=_re.IGNORECASE,
                )
                break

    # ── IVA manual ────────────────────────────────────────────────────────────
    if cfg.get("iva_manual"):
        ws[cfg["iva_manual"]] = cotizacion.valor_iva or 0

    # ── Página ───────────────────────────────────────────────────────────────
    print_area = empresa.get("print_area")
    if print_area:
        ws.print_area = print_area

    # Eliminar columnas espurias más allá de la columna J
    max_content_col = 10
    for col in [c for c in list(ws.column_dimensions.keys())
                if column_index_from_string(c) > max_content_col]:
        del ws.column_dimensions[col]

    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = "landscape" if empresa.get("orientacion_paisaje") else "portrait"
    ws.page_setup.paperSize   = 9  # A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75, header=0.3, footer=0.3)

    wb.save(str(output_path))


def _intentar_conversion_pdf(source_path: Path, output_dir: Path) -> Optional[Path]:
    """Convierte un XLSX (o DOCX) a PDF con LibreOffice. Retorna la ruta del PDF si tiene éxito."""
    import os
    import logging
    _log = logging.getLogger(__name__)
    try:
        # -env:UserInstallation es la forma correcta en LibreOffice 25.x de
        # apuntar el perfil a un dir escribible (--user-data-dir fue deprecado).
        # HOME=/tmp evita fallos de permisos en contenedores Docker.
        env = {**os.environ, "HOME": "/tmp"}
        result = subprocess.run(
            [
                "libreoffice",
                "--headless",
                "-env:UserInstallation=file:///tmp/lo_userdata",
                "--convert-to",
                "pdf",
                "--outdir",
                str(output_dir),
                str(source_path),
            ],
            timeout=90,
            capture_output=True,
            env=env,
        )
        if result.returncode == 0:
            pdf_path = output_dir / (source_path.stem + ".pdf")
            if pdf_path.exists():
                return pdf_path
            _log.warning("LibreOffice exit 0 pero PDF no encontrado: %s", pdf_path)
        else:
            _log.warning(
                "LibreOffice falló (rc=%d): %s",
                result.returncode,
                result.stderr.decode(errors="replace")[:500],
            )
    except FileNotFoundError:
        _log.warning("LibreOffice no encontrado en el sistema.")
    except subprocess.TimeoutExpired:
        _log.warning("LibreOffice timeout al convertir %s", source_path)
    return None


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.post(
    "/solicitudes/{solicitud_id}/generar-oc",
    response_model=OrdenCompraRead,
    status_code=status.HTTP_201_CREATED,
)
def generar_orden_compra(
    solicitud_id: uuid.UUID,
    forzar: bool = False,
    current_user: User = Depends(require_compras),
    oc_db: Session = Depends(get_oc_db),
    db: Session = Depends(get_db),
):
    # Verificar si ya existe una OC para esta solicitud
    orden_existente = oc_db.exec(
        select(OrdenCompra).where(OrdenCompra.solicitud_id == solicitud_id)
    ).first()

    if orden_existente and not forzar:
        return orden_existente

    # Buscar la solicitud
    solicitud = oc_db.get(SolicitudOC, solicitud_id)
    if not solicitud:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Solicitud no encontrada.",
        )

    # Buscar cotización aprobada más reciente
    cotizacion = oc_db.exec(
        select(CotizacionProveedor)
        .where(
            CotizacionProveedor.solicitud_id == solicitud_id,
            CotizacionProveedor.aprobada == True,  # noqa: E712
        )
        .order_by(CotizacionProveedor.created_at.desc())
    ).first()
    if not cotizacion:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No existe una cotización aprobada para esta solicitud.",
        )

    # Si es regeneración, reutilizar el número existente; si no, generar nuevo
    if orden_existente and forzar:
        numero_oc = orden_existente.numero_oc
        # Eliminar archivos viejos para que no queden huérfanos
        for viejo in [
            OC_DOCS_DIR / f"{numero_oc}.xlsx",
            OC_DOCS_DIR / f"{numero_oc}.pdf",
        ]:
            if viejo.exists():
                viejo.unlink(missing_ok=True)
    else:
        # Generar número de OC: OC-{año}-{secuencial 4 dígitos}
        anio_actual = datetime.now(timezone.utc).year
        prefijo = f"OC-{anio_actual}-"
        count = oc_db.exec(
            select(func.count(OrdenCompra.id)).where(
                OrdenCompra.numero_oc.startswith(prefijo)
            )
        ).one()
        numero_oc = f"{prefijo}{(count + 1):04d}"

    # Preparar directorio de salida
    OC_DOCS_DIR.mkdir(parents=True, exist_ok=True)
    xlsx_path = OC_DOCS_DIR / f"{numero_oc}.xlsx"

    # Resolver nombres para firmas
    auxiliar_nombre = ""
    aprobador_nombre = ""
    if solicitud.auxiliar_id:
        aux = db.get(User, solicitud.auxiliar_id)
        if aux:
            auxiliar_nombre = aux.full_name
    if cotizacion.aprobado_por_id:
        aprobador = db.get(User, cotizacion.aprobado_por_id)
        if aprobador:
            aprobador_nombre = aprobador.full_name

    # Generar XLSX desde plantilla
    _generar_xlsx(numero_oc, solicitud, cotizacion, xlsx_path, auxiliar_nombre, aprobador_nombre)

    # Intentar conversión a PDF (LibreOffice recalcula fórmulas)
    pdf_path_obj = _intentar_conversion_pdf(xlsx_path, OC_DOCS_DIR)
    pdf_path_str = str(pdf_path_obj) if pdf_path_obj else None

    # Actualizar registro existente o crear uno nuevo
    if orden_existente and forzar:
        orden_existente.cotizacion_id = cotizacion.id
        orden_existente.pdf_path = pdf_path_str
        oc_db.add(orden_existente)
        oc_db.commit()
        oc_db.refresh(orden_existente)
        return orden_existente

    orden = OrdenCompra(
        solicitud_id=solicitud_id,
        cotizacion_id=cotizacion.id,
        numero_oc=numero_oc,
        pdf_path=pdf_path_str,
        email_proveedor=cotizacion.proveedor_email,
        created_at=datetime.now(timezone.utc),
    )
    oc_db.add(orden)
    oc_db.commit()
    oc_db.refresh(orden)
    return orden


@router.get(
    "/solicitudes/{solicitud_id}/orden",
    response_model=OrdenCompraRead,
)
def obtener_orden_por_solicitud(
    solicitud_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    oc_db: Session = Depends(get_oc_db),
):
    orden = oc_db.exec(
        select(OrdenCompra).where(OrdenCompra.solicitud_id == solicitud_id)
    ).first()
    if not orden:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No existe orden de compra para esta solicitud.",
        )
    return orden


@router.post(
    "/solicitudes/{solicitud_id}/marcar-enviada",
    response_model=None,
    status_code=status.HTTP_200_OK,
)
async def marcar_oc_enviada(
    solicitud_id: uuid.UUID,
    payload: MarcarEnviadaPayload,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(require_compras),
    oc_db: Session = Depends(get_oc_db),
):
    from app.models.oc import EstadoOC
    from app.services import email_service

    solicitud = oc_db.get(SolicitudOC, solicitud_id)
    if not solicitud:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Solicitud no encontrada.")
    if solicitud.estado != EstadoOC.aprobada:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Solo se puede marcar como enviada desde estado 'aprobada'. Estado actual: {solicitud.estado}",
        )

    orden = oc_db.exec(
        select(OrdenCompra).where(OrdenCompra.solicitud_id == solicitud_id)
    ).first()

    estado_anterior = solicitud.estado
    solicitud.estado = EstadoOC.oc_enviada
    solicitud.fecha_envio_oc = datetime.now(timezone.utc)
    solicitud.updated_at = datetime.now(timezone.utc)
    oc_db.add(solicitud)

    registrar_cambio_estado(
        oc_db,
        solicitud.id,
        estado_anterior,
        EstadoOC.oc_enviada,
        usuario_id=current_user.id,
        usuario_nombre=current_user.full_name,
    )

    if orden and payload.email_proveedor:
        orden.email_proveedor = payload.email_proveedor
        orden.enviada_proveedor = True
        oc_db.add(orden)

    oc_db.commit()
    oc_db.refresh(solicitud)

    background_tasks.add_task(email_service.send_oc_enviada, solicitud)

    if orden and payload.email_proveedor:
        # Buscar ítems de la cotización aprobada para enviarlos al proveedor
        cotizacion = oc_db.exec(
            select(CotizacionProveedor)
            .where(
                CotizacionProveedor.solicitud_id == solicitud_id,
                CotizacionProveedor.aprobada == True,  # noqa: E712
            )
            .order_by(CotizacionProveedor.created_at.desc())
        ).first()
        items_cot = cotizacion.items if cotizacion else None

        background_tasks.add_task(
            email_service.send_oc_a_proveedor,
            solicitud,
            orden.numero_oc,
            orden.pdf_path,
            payload.email_proveedor,
            items_cot,
        )

    return {"ok": True}


@router.post(
    "/solicitudes/{solicitud_id}/marcar-en-plataforma",
    response_model=None,
    status_code=status.HTTP_200_OK,
)
def marcar_en_plataforma(
    solicitud_id: uuid.UUID,
    current_user: User = Depends(require_compras),
    oc_db: Session = Depends(get_oc_db),
):
    """Auxiliar confirma que el pedido ya fue ingresado en la plataforma (ERP/sistema)."""
    from app.models.oc import EstadoOC

    solicitud = oc_db.get(SolicitudOC, solicitud_id)
    if not solicitud:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Solicitud no encontrada.")
    if solicitud.estado != EstadoOC.oc_enviada:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Solo se puede marcar en plataforma desde estado 'oc_enviada'. Estado actual: {solicitud.estado}",
        )

    estado_anterior = solicitud.estado
    solicitud.estado = EstadoOC.oc_en_plataforma
    solicitud.fecha_en_plataforma = datetime.now(timezone.utc)
    solicitud.updated_at = datetime.now(timezone.utc)
    oc_db.add(solicitud)

    registrar_cambio_estado(
        oc_db,
        solicitud.id,
        estado_anterior,
        EstadoOC.oc_en_plataforma,
        usuario_id=current_user.id,
        usuario_nombre=current_user.full_name,
    )

    oc_db.commit()

    return {"ok": True}


@router.post(
    "/solicitudes/{solicitud_id}/marcar-entregada",
    response_model=None,
    status_code=status.HTTP_200_OK,
)
def marcar_entregada(
    solicitud_id: uuid.UUID,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(get_current_user),
    oc_db: Session = Depends(get_oc_db),
):
    """Coordinador/solicitante confirma que recibió físicamente el pedido.
    Accesible para cualquier usuario autenticado cuyo email coincida con el solicitante,
    o para usuarios del módulo de compras (require_compras)."""
    from app.models.oc import EstadoOC
    from app.services import email_service

    OC_ROLES = {"admin", "administrativo", "directivo", "compras"}
    es_compras = current_user.role in OC_ROLES or current_user.area == "Compras"

    solicitud = oc_db.get(SolicitudOC, solicitud_id)
    if not solicitud:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Solicitud no encontrada.")

    # Validar que sea el solicitante o alguien de compras
    if not es_compras and solicitud.solicitante_email != current_user.email:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Solo el solicitante o el equipo de compras puede confirmar la recepción.",
        )

    if solicitud.estado != EstadoOC.oc_en_plataforma:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Solo se puede confirmar recepción desde estado 'oc_en_plataforma'. Estado actual: {solicitud.estado}",
        )

    estado_anterior = solicitud.estado
    solicitud.estado = EstadoOC.entregada
    solicitud.fecha_recibido = datetime.now(timezone.utc)
    solicitud.updated_at = datetime.now(timezone.utc)
    oc_db.add(solicitud)

    registrar_cambio_estado(
        oc_db,
        solicitud.id,
        estado_anterior,
        EstadoOC.entregada,
        usuario_id=current_user.id,
        usuario_nombre=current_user.full_name,
    )

    oc_db.commit()
    oc_db.refresh(solicitud)

    background_tasks.add_task(email_service.send_entrega_confirmada, solicitud)

    return {"ok": True}


@router.post(
    "/solicitudes/{solicitud_id}/cerrar",
    response_model=None,
    status_code=status.HTTP_200_OK,
)
def cerrar_solicitud(
    solicitud_id: uuid.UUID,
    current_user: User = Depends(require_compras),
    oc_db: Session = Depends(get_oc_db),
):
    from app.models.oc import EstadoOC

    solicitud = oc_db.get(SolicitudOC, solicitud_id)
    if not solicitud:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Solicitud no encontrada.")
    if solicitud.estado != EstadoOC.entregada:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Solo se puede cerrar desde estado 'entregada'. Estado actual: {solicitud.estado}",
        )

    estado_anterior = solicitud.estado
    solicitud.estado = EstadoOC.cerrada
    solicitud.updated_at = datetime.now(timezone.utc)
    oc_db.add(solicitud)

    registrar_cambio_estado(
        oc_db,
        solicitud.id,
        estado_anterior,
        EstadoOC.cerrada,
        usuario_id=current_user.id,
        usuario_nombre=current_user.full_name,
    )

    oc_db.commit()

    return {"ok": True}


@router.get("/ordenes/{orden_id}/descargar")
def descargar_orden(
    orden_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    oc_db: Session = Depends(get_oc_db),
):
    orden = oc_db.get(OrdenCompra, orden_id)
    if not orden:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Orden de compra no encontrada.",
        )

    # Intentar servir PDF si existe
    if orden.pdf_path:
        pdf_path = Path(orden.pdf_path)
        if pdf_path.exists():
            return FileResponse(
                path=str(pdf_path),
                media_type="application/pdf",
                filename=f"{orden.numero_oc}.pdf",
            )

    # Fallback: servir XLSX si el PDF no está disponible
    xlsx_path = OC_DOCS_DIR / f"{orden.numero_oc}.xlsx"
    if xlsx_path.exists():
        return FileResponse(
            path=str(xlsx_path),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=f"{orden.numero_oc}.xlsx",
        )

    raise HTTPException(
        status_code=status.HTTP_404_NOT_FOUND,
        detail="No se encontró el archivo de la orden de compra.",
    )
