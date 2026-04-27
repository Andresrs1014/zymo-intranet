import io
import json
import logging
import random
import smtplib
import ssl
import zipfile
from datetime import date, timedelta
from typing import Optional

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlmodel import Session, select

from app.config import settings
from app.core.deps import get_current_user
from app.models.oc import OcConfig
from app.models.user import User
from app.oc_database import get_oc_db
from app.services.field_synonyms import FIELD_SYNONYMS

log = logging.getLogger(__name__)

router = APIRouter(tags=["OC - Configuración"])

_ALLOWED_KEYS = {
    # SMTP
    "smtp_host", "smtp_port", "smtp_user", "smtp_password", "smtp_from",
    # Destinatarios
    "email_directora", "email_compras", "email_contabilidad", "intranet_url",
    # Branding por plataforma
    "empresa_nombre", "empresa_color", "empresa_nit", "empresa_tel", "empresa_dir", "empresa_dept",
    # Asunto y plantillas de email
    "email_prefijo",
    "email_intro_flujo1", "email_intro_flujo2", "email_intro_flujo3", "email_intro_flujo4",
}


def _require_admin(user: User) -> None:
    if user.role != "admin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Solo administradores.")


# ── Schemas ───────────────────────────────────────────────────────────────────

class ConfigRead(BaseModel):
    smtp_host: str
    smtp_port: str
    smtp_user: str
    smtp_password_set: bool
    smtp_from: str
    email_directora: str
    email_compras: str
    email_contabilidad: str
    intranet_url: str
    # Branding
    empresa_nombre: str
    empresa_color: str
    empresa_nit: str
    empresa_tel: str
    empresa_dir: str
    empresa_dept: str
    # Email templates
    email_prefijo: str
    email_intro_flujo1: str
    email_intro_flujo2: str
    email_intro_flujo3: str
    email_intro_flujo4: str


class ConfigUpdate(BaseModel):
    smtp_host: Optional[str] = None
    smtp_port: Optional[str] = None
    smtp_user: Optional[str] = None
    smtp_password: Optional[str] = None
    smtp_from: Optional[str] = None
    email_directora: Optional[str] = None
    email_compras: Optional[str] = None
    email_contabilidad: Optional[str] = None
    intranet_url: Optional[str] = None
    # Branding
    empresa_nombre: Optional[str] = None
    empresa_color: Optional[str] = None
    empresa_nit: Optional[str] = None
    empresa_tel: Optional[str] = None
    empresa_dir: Optional[str] = None
    empresa_dept: Optional[str] = None
    # Email templates
    email_prefijo: Optional[str] = None
    email_intro_flujo1: Optional[str] = None
    email_intro_flujo2: Optional[str] = None
    email_intro_flujo3: Optional[str] = None
    email_intro_flujo4: Optional[str] = None


class TestEmailResult(BaseModel):
    ok: bool
    mensaje: str
    detalle: Optional[str] = None


class ListasFormulario(BaseModel):
    prioridades: list[str]
    categorias: list[str]
    grupos_articulos: list[str]
    clientes: list[str]
    condiciones: list[str]
    placas: list[str]


_DEFAULT_LISTAS: dict[str, list[str]] = {
    "lista_prioridades":      ["Alta", "Media", "Baja"],
    "lista_categorias":       ["Repuesto", "Insumo", "Herramienta", "Servicio", "Otro"],
    "lista_grupos_articulos": ["Mecánico", "Eléctrico", "Hidráulico", "Neumático", "Lubricantes", "Papelería", "Otro"],
    "lista_clientes":         [],
    "lista_condiciones":      ["Nuevo", "Reposición", "Urgente"],
    "lista_placas":           [],
}


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.get("/config", response_model=ConfigRead)
def get_config(
    current_user: User = Depends(get_current_user),
    oc_db: Session = Depends(get_oc_db),
):
    _require_admin(current_user)
    rows = {r.key: r.value for r in oc_db.exec(select(OcConfig)).all()}

    from app.services.email_service import _BRANDING_DEFAULTS

    def _r(key: str, fallback: str = "") -> str:
        return rows.get(key) or fallback

    return ConfigRead(
        smtp_host=_r("smtp_host", settings.smtp_host),
        smtp_port=_r("smtp_port", str(settings.smtp_port)),
        smtp_user=_r("smtp_user", settings.smtp_user),
        smtp_password_set=bool(rows.get("smtp_password") or settings.smtp_password),
        smtp_from=_r("smtp_from", settings.smtp_from or settings.smtp_user),
        email_directora=_r("email_directora", settings.email_directora),
        email_compras=_r("email_compras", settings.email_directora),
        email_contabilidad=_r("email_contabilidad"),
        intranet_url=_r("intranet_url", settings.intranet_url),
        # Branding — usa _BRANDING_DEFAULTS cuando no hay valor en DB
        empresa_nombre=_r("empresa_nombre", _BRANDING_DEFAULTS["empresa_nombre"]),
        empresa_color=_r("empresa_color", _BRANDING_DEFAULTS["empresa_color"]),
        empresa_nit=_r("empresa_nit", _BRANDING_DEFAULTS["empresa_nit"]),
        empresa_tel=_r("empresa_tel", _BRANDING_DEFAULTS["empresa_tel"]),
        empresa_dir=_r("empresa_dir", _BRANDING_DEFAULTS["empresa_dir"]),
        empresa_dept=_r("empresa_dept", _BRANDING_DEFAULTS["empresa_dept"]),
        # Email templates — vacío = el email_service usa _INTRO_DEFAULTS
        email_prefijo=_r("email_prefijo", _BRANDING_DEFAULTS["email_prefijo"]),
        email_intro_flujo1=_r("email_intro_flujo1", ""),
        email_intro_flujo2=_r("email_intro_flujo2", ""),
        email_intro_flujo3=_r("email_intro_flujo3", ""),
        email_intro_flujo4=_r("email_intro_flujo4", ""),
    )


@router.patch("/config", status_code=status.HTTP_204_NO_CONTENT)
def update_config(
    payload: ConfigUpdate,
    current_user: User = Depends(get_current_user),
    oc_db: Session = Depends(get_oc_db),
):
    _require_admin(current_user)

    for field, value in payload.model_dump(exclude_unset=True).items():
        if value is None or field not in _ALLOWED_KEYS:
            continue
        existing = oc_db.get(OcConfig, field)
        if existing:
            existing.value = value
            oc_db.add(existing)
        else:
            oc_db.add(OcConfig(key=field, value=value))

    oc_db.commit()


@router.get("/config/listas", response_model=ListasFormulario)
def get_listas(
    current_user: User = Depends(get_current_user),
    oc_db: Session = Depends(get_oc_db),
) -> ListasFormulario:
    """Retorna las listas configurables del formulario de solicitud.
    Accesible para cualquier usuario autenticado (el formulario las necesita)."""
    rows = {r.key: r.value for r in oc_db.exec(select(OcConfig)).all()}

    def _get_lista(key: str) -> list[str]:
        raw = rows.get(key)
        if raw:
            try:
                return json.loads(raw)
            except Exception:
                return []
        return _DEFAULT_LISTAS.get(key, [])

    return ListasFormulario(
        prioridades=_get_lista("lista_prioridades"),
        categorias=_get_lista("lista_categorias"),
        grupos_articulos=_get_lista("lista_grupos_articulos"),
        clientes=_get_lista("lista_clientes"),
        condiciones=_get_lista("lista_condiciones"),
        placas=_get_lista("lista_placas"),
    )


@router.patch("/config/listas", status_code=status.HTTP_204_NO_CONTENT)
def update_listas(
    payload: ListasFormulario,
    current_user: User = Depends(get_current_user),
    oc_db: Session = Depends(get_oc_db),
) -> None:
    """Guarda las listas configurables. Solo admin."""
    _require_admin(current_user)

    mapping = {
        "lista_prioridades":      payload.prioridades,
        "lista_categorias":       payload.categorias,
        "lista_grupos_articulos": payload.grupos_articulos,
        "lista_clientes":         payload.clientes,
        "lista_condiciones":      payload.condiciones,
        "lista_placas":           payload.placas,
    }
    for key, value in mapping.items():
        serialized = json.dumps(value, ensure_ascii=False)
        existing = oc_db.get(OcConfig, key)
        if existing:
            existing.value = serialized
            oc_db.add(existing)
        else:
            oc_db.add(OcConfig(key=key, value=serialized))

    oc_db.commit()


@router.post("/config/listas/upload-clientes", response_model=ListasFormulario)
async def upload_clientes_excel(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    oc_db: Session = Depends(get_oc_db),
):
    """Sube un Excel de clientes, lee la primera columna y actualiza la lista_clientes. Solo admin."""
    _require_admin(current_user)

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Formato no soportado. Sube un archivo .xlsx o .xls",
        )

    try:
        import openpyxl
        import io

        contenido = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
        ws = wb.active
        nuevos_clientes = set()

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                cliente = str(row[0]).strip()
                if cliente:
                    nuevos_clientes.add(cliente)

        lista_clientes = sorted(list(nuevos_clientes))

        # Guardar en base de datos
        serialized = json.dumps(lista_clientes, ensure_ascii=False)
        existing = oc_db.get(OcConfig, "lista_clientes")
        if existing:
            existing.value = serialized
            oc_db.add(existing)
        else:
            oc_db.add(OcConfig(key="lista_clientes", value=serialized))

        oc_db.commit()

        # Retornar las listas actualizadas
        return get_listas(current_user, oc_db)

    except Exception as e:
        log.warning("[config] Error al leer excel de clientes: %s", e)
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"Error al procesar el archivo Excel: {e}",
        )


@router.post("/config/test-email", response_model=TestEmailResult)
def test_email(
    current_user: User = Depends(get_current_user),
    oc_db: Session = Depends(get_oc_db),
):
    """Envía un correo de prueba al smtp_user para verificar que las credenciales funcionan.
    Usa smtplib directamente para obtener el error SMTP exacto sin capturarlo en silencio.
    """
    _require_admin(current_user)

    rows = {r.key: r.value for r in oc_db.exec(select(OcConfig)).all()}
    host = rows.get("smtp_host") or settings.smtp_host
    port = int(rows.get("smtp_port") or settings.smtp_port)
    user = rows.get("smtp_user") or settings.smtp_user
    password = rows.get("smtp_password") or settings.smtp_password
    from_addr = rows.get("smtp_from") or settings.smtp_from or user
    to_addr = user  # el correo de prueba se manda al mismo remitente

    if not user or not password:
        return TestEmailResult(
            ok=False,
            mensaje="Credenciales SMTP no configuradas.",
            detalle="smtp_user y smtp_password deben estar configurados antes de hacer la prueba.",
        )

    # Construir mensaje mínimo (sin fastapi-mail para ver el error real de smtplib)
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart

    msg = MIMEMultipart("alternative")
    msg["Subject"] = "[ZYMO Intranet] Correo de prueba SMTP"
    msg["From"] = f"Compras ZYMO <{from_addr}>"
    msg["To"] = to_addr
    html = """
    <div style="font-family:Arial,sans-serif;padding:24px;max-width:500px">
      <h3 style="color:#111">✅ Prueba SMTP exitosa</h3>
      <p style="color:#444">Este correo confirma que la configuración SMTP de ZYMO Intranet
      está funcionando correctamente.</p>
      <p style="color:#888;font-size:12px">Generado desde /oc/configuracion</p>
    </div>
    """
    msg.attach(MIMEText(html, "html", "utf-8"))

    try:
        context = ssl.create_default_context()
        with smtplib.SMTP(host, port, timeout=15) as smtp:
            smtp.ehlo()
            smtp.starttls(context=context)
            smtp.ehlo()
            smtp.login(user, password)
            smtp.sendmail(from_addr, [to_addr], msg.as_string())

        log.info("[email_test] Prueba SMTP exitosa → %s", to_addr)
        return TestEmailResult(
            ok=True,
            mensaje=f"Correo enviado correctamente a {to_addr}.",
            detalle=f"Servidor: {host}:{port} | Usuario: {user}",
        )

    except smtplib.SMTPAuthenticationError as e:
        detalle = (
            "Error de autenticación. Causas frecuentes en Office 365:\n"
            "1. SMTP AUTH deshabilitado en el buzón (habilitar en Centro de administración M365 → Usuarios → Correo → Autenticación SMTP).\n"
            "2. MFA activa sin contraseña de aplicación (crear una en Seguridad de la cuenta).\n"
            "3. Usuario o contraseña incorrectos.\n"
            f"Código SMTP: {e.smtp_code} — {e.smtp_error!r}"
        )
        log.warning("[email_test] SMTPAuthenticationError: %s %r", e.smtp_code, e.smtp_error)
        return TestEmailResult(ok=False, mensaje="Fallo de autenticación SMTP.", detalle=detalle)

    except smtplib.SMTPConnectError as e:
        detalle = (
            f"No se pudo conectar a {host}:{port}.\n"
            "Verifica que el host y puerto sean correctos y que el servidor esté accesible desde Docker.\n"
            f"Error: {e}"
        )
        log.warning("[email_test] SMTPConnectError: %s", e)
        return TestEmailResult(ok=False, mensaje="No se pudo conectar al servidor SMTP.", detalle=detalle)

    except smtplib.SMTPException as e:
        log.warning("[email_test] SMTPException: %s", e)
        return TestEmailResult(ok=False, mensaje="Error SMTP.", detalle=str(e))

    except TimeoutError:
        detalle = (
            f"Timeout conectando a {host}:{port} (15 s).\n"
            "El servidor no responde. Posible bloqueo de firewall o puerto incorrecto."
        )
        log.warning("[email_test] Timeout conectando a %s:%s", host, port)
        return TestEmailResult(ok=False, mensaje="Timeout de conexión.", detalle=detalle)

    except Exception as e:
        log.exception("[email_test] Error inesperado")
        return TestEmailResult(ok=False, mensaje="Error inesperado.", detalle=str(e))


# ── Test Excel generator ──────────────────────────────────────────────────────

_EMPRESAS = [
    "Suministros del Norte S.A.S.", "Proveedores Industriales Ltda.", "Papelería El Centro S.A.",
    "Distribuidora Tech Colombia S.A.S.", "Insumos y Repuestos del Valle Ltda.",
    "Ferretería Industrial Bogotá S.A.", "Oficina Total S.A.S.", "Comercializadora Andina Ltda.",
    "Electrónica y Sistemas S.A.", "Materiales y Construcción del Sur S.A.S.",
]

_ITEMS_POOL = [
    ("Resma de papel carta 75g", "PAP-001", 10, 12500),
    ("Tóner HP LaserJet 85A", "TON-HP85A", 3, 89000),
    ("Bolígrafo BIC punto fino azul x12", "ESC-BIC12", 5, 8900),
    ("Carpeta de palanca tamaño carta", "ARC-CAR01", 20, 5200),
    ("Marcador permanente Sharpie negro", "MAR-SHP01", 8, 3500),
    ("Cinta adhesiva transparente 48mm", "CIN-48T", 12, 2800),
    ("Grapadora metálica 26/6", "GRA-MET26", 4, 18500),
    ("Resaltador fluorescente amarillo", "RES-AMA01", 15, 1900),
    ("Sobre manila carta x100", "SOB-MAN100", 6, 14200),
    ("Perforadora 2 huecos metálica", "PER-2H01", 3, 22000),
    ("Sello automático personalizado", "SEL-AUT01", 2, 45000),
    ("Post-it notas adhesivas 76x76mm", "NOT-POST76", 10, 7500),
    ("Tijeras de oficina 20cm", "TIJ-20CM", 5, 6800),
    ("Regla metálica 30cm", "REG-30M", 8, 4500),
    ("Lapicero gel 0.5mm negro x10", "LAP-GEL10", 4, 12000),
    ("Borrador blanco para lápiz", "BOR-BLA01", 20, 800),
    ("Clip mariposa mediano x50", "CLI-MAR50", 10, 3200),
    ("Folder plástico oficio transparente", "FOL-PLA01", 15, 2100),
    ("Corrector líquido tipo lapicero", "COR-LIQ01", 6, 4200),
    ("Caja de clips estándar x100", "CLI-EST100", 12, 2500),
]

_FORMAS_PAGO = ["Contado", "Crédito 30 días", "Crédito 15 días", "50% anticipo - 50% contra entrega", "Transferencia inmediata"]
_PLAZOS = ["Inmediata", "3 días hábiles", "5 días hábiles", "8 días hábiles", "15 días calendario"]
_GARANTIAS = ["6 meses", "1 año fabricante", "30 días por defectos de fábrica", "No aplica", "12 meses"]
_ANTICIPOS = ["No requiere", "30%", "50%", "100% anticipado"]
_SALDOS = ["Contra entrega", "Al facturar", "30 días", "No aplica"]


def _rnd_syn(campo: str) -> str:
    """Retorna un sinónimo aleatorio del campo (incluye el propio nombre canónico)."""
    opciones = [campo] + FIELD_SYNONYMS.get(campo, [])
    return random.choice(opciones)


def _generar_excel_prueba() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cotización"

    empresa = random.choice(_EMPRESAS)
    nit = f"{random.randint(800, 999)}.{random.randint(100, 999)}.{random.randint(100, 999)}-{random.randint(0, 9)}"
    email = f"ventas@{empresa.split()[0].lower().replace(',', '')}.com"
    num_cot = f"COT-{random.randint(2025,2026)}-{random.randint(100,999)}"
    vigencia = (date.today() + timedelta(days=random.randint(10, 30))).isoformat()

    # ── Encabezado escalar ────────────────────────────────────────────────────
    fila = 1
    ws.cell(fila, 1, "COTIZACIÓN DE COMPRA")
    fila += 1
    ws.cell(fila, 1, _rnd_syn("numero_cotizacion_proveedor").title())
    ws.cell(fila, 2, num_cot)
    fila += 1
    ws.cell(fila, 1, _rnd_syn("proveedor_nombre").title())
    ws.cell(fila, 2, empresa)
    fila += 1
    ws.cell(fila, 1, _rnd_syn("proveedor_nit").upper())
    ws.cell(fila, 2, nit)
    fila += 1
    ws.cell(fila, 1, _rnd_syn("proveedor_email").title())
    ws.cell(fila, 2, email)
    fila += 1
    ws.cell(fila, 1, _rnd_syn("fecha_vigencia").title())
    ws.cell(fila, 2, vigencia)
    fila += 1
    ws.cell(fila, 1, _rnd_syn("forma_pago").title())
    ws.cell(fila, 2, random.choice(_FORMAS_PAGO))
    fila += 1
    ws.cell(fila, 1, _rnd_syn("plazo_entrega").title())
    ws.cell(fila, 2, random.choice(_PLAZOS))
    fila += 1
    ws.cell(fila, 1, _rnd_syn("garantia").title())
    ws.cell(fila, 2, random.choice(_GARANTIAS))
    fila += 1
    ws.cell(fila, 1, _rnd_syn("anticipo").title())
    ws.cell(fila, 2, random.choice(_ANTICIPOS))
    fila += 1
    ws.cell(fila, 1, _rnd_syn("pago_saldo").title())
    ws.cell(fila, 2, random.choice(_SALDOS))
    fila += 2  # fila en blanco

    # ── Tabla de ítems ────────────────────────────────────────────────────────
    col_desc = _rnd_syn("descripcion").title()
    col_ref  = _rnd_syn("referencia").title()
    col_cant = _rnd_syn("cantidad").title()
    col_unit = _rnd_syn("valor_unitario").title()
    col_tot  = _rnd_syn("valor_total").title()

    ws.cell(fila, 1, "No.")
    ws.cell(fila, 2, col_desc)
    ws.cell(fila, 3, col_ref)
    ws.cell(fila, 4, col_cant)
    ws.cell(fila, 5, col_unit)
    ws.cell(fila, 6, col_tot)
    fila += 1

    n_items = random.randint(3, 10)
    pool = random.sample(_ITEMS_POOL, min(n_items, len(_ITEMS_POOL)))
    subtotal = 0.0

    for i, (desc, ref, cant_base, precio_base) in enumerate(pool, 1):
        cant = random.randint(1, cant_base)
        precio = round(precio_base * random.uniform(0.9, 1.15), -2)  # variación ±15%
        total_fila = round(cant * precio, 2)
        subtotal += total_fila
        ws.cell(fila, 1, i)
        ws.cell(fila, 2, desc)
        ws.cell(fila, 3, ref)
        ws.cell(fila, 4, cant)
        ws.cell(fila, 5, precio)
        ws.cell(fila, 6, total_fila)
        fila += 1

    fila += 1  # fila en blanco

    iva = round(subtotal * 0.19, 2)
    total = round(subtotal + iva, 2)

    ws.cell(fila, 5, _rnd_syn("valor_antes_iva").title())
    ws.cell(fila, 6, subtotal)
    fila += 1
    ws.cell(fila, 5, _rnd_syn("valor_iva").upper())
    ws.cell(fila, 6, iva)
    fila += 1
    ws.cell(fila, 5, _rnd_syn("valor_total").title())
    ws.cell(fila, 6, total)

    # Ajustar ancho columnas
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@router.get("/config/test/generar-excel")
def generar_excel_prueba(
    current_user: User = Depends(get_current_user),
    oc_db: Session = Depends(get_oc_db),
):
    """Genera un Excel de prueba con datos aleatorios usando sinónimos del motor."""
    _require_admin(current_user)

    counter_row = oc_db.get(OcConfig, "test_excel_counter")
    counter = int(counter_row.value) + 1 if counter_row else 1
    if counter_row:
        counter_row.value = str(counter)
        oc_db.add(counter_row)
    else:
        oc_db.add(OcConfig(key="test_excel_counter", value=str(counter)))
    oc_db.commit()

    filename = f"prueba.{counter:03d}.xlsx"
    contenido = _generar_excel_prueba()

    return StreamingResponse(
        io.BytesIO(contenido),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/config/test/generar-par")
def generar_par_prueba(
    current_user: User = Depends(get_current_user),
    oc_db: Session = Depends(get_oc_db),
):
    """Genera un ZIP con una cotización Excel + una factura PDF colombiana con los mismos datos.
    Útil para probar el motor de validación financiera end-to-end."""
    _require_admin(current_user)

    counter_row = oc_db.get(OcConfig, "test_par_counter")
    counter = int(counter_row.value) + 1 if counter_row else 1
    if counter_row:
        counter_row.value = str(counter)
        oc_db.add(counter_row)
    else:
        oc_db.add(OcConfig(key="test_par_counter", value=str(counter)))
    oc_db.commit()

    datos = _generar_datos_prueba()
    excel_bytes = _excel_desde_datos(datos)
    pdf_bytes = _factura_pdf_desde_datos(datos)

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"cotizacion.{counter:03d}.xlsx", excel_bytes)
        zf.writestr(f"factura.{counter:03d}.pdf", pdf_bytes)
    zip_buf.seek(0)

    return StreamingResponse(
        zip_buf,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="par-prueba.{counter:03d}.zip"'},
    )


# ── Generadores internos ──────────────────────────────────────────────────────

_CIUDADES_CO = ["Bogotá D.C.", "Medellín", "Cali", "Barranquilla", "Bucaramanga", "Pereira", "Manizales"]
_CALLES_CO = ["Cra. 7", "Calle 72", "Av. El Dorado", "Cra. 13", "Calle 100", "Av. 68", "Cra. 30"]

_ADQUIRENTE = {
    "nombre": "Zymo Logística S.A.S.",
    "nit": "900.123.456-7",
    "direccion": "Calle 25 # 85-35, Bodega 3",
    "ciudad": "Bogotá D.C.",
    "email": "compras@zymo.com",
    "regimen": "Responsable de IVA",
}


def _generar_datos_prueba() -> dict:
    """Genera un set de datos aleatorio consistente para cotización y factura."""
    empresa = random.choice(_EMPRESAS)
    nit = f"{random.randint(800, 999)}.{random.randint(100, 999)}.{random.randint(100, 999)}-{random.randint(0, 9)}"
    email = f"ventas@{empresa.split()[0].lower().replace(',', '').replace('.', '')}.com"
    num_cot = f"COT-{date.today().year}-{random.randint(100, 999)}"
    num_fact = f"FACT-{date.today().year}-{random.randint(1000, 9999)}"
    vigencia = (date.today() + timedelta(days=random.randint(10, 30))).isoformat()
    ciudad = random.choice(_CIUDADES_CO)
    calle = random.choice(_CALLES_CO)
    direccion = f"{calle} # {random.randint(1,99)}-{random.randint(10,99)}"
    forma_pago = random.choice(_FORMAS_PAGO)
    plazo = random.choice(_PLAZOS)
    garantia = random.choice(_GARANTIAS)
    anticipo = random.choice(_ANTICIPOS)
    saldo = random.choice(_SALDOS)
    resolucion_dian = f"{random.randint(18000000, 18999999)}"
    res_desde = random.randint(1000, 4999)
    res_hasta = res_desde + random.randint(1000, 5000)
    res_fecha = (date.today() - timedelta(days=random.randint(30, 365))).isoformat()

    n_items = random.randint(3, 8)
    pool = random.sample(_ITEMS_POOL, min(n_items, len(_ITEMS_POOL)))
    items = []
    subtotal = 0.0
    for i, (desc, ref, cant_base, precio_base) in enumerate(pool, 1):
        cant = random.randint(1, cant_base)
        precio = round(precio_base * random.uniform(0.9, 1.15), -2)
        total_fila = round(cant * precio, 2)
        subtotal += total_fila
        items.append({"num": i, "desc": desc, "ref": ref, "cant": cant, "precio": precio, "total": total_fila})

    iva = round(subtotal * 0.19, 2)
    total = round(subtotal + iva, 2)

    return {
        "empresa": empresa, "nit": nit, "email": email, "direccion": direccion, "ciudad": ciudad,
        "num_cot": num_cot, "num_fact": num_fact, "vigencia": vigencia,
        "forma_pago": forma_pago, "plazo": plazo, "garantia": garantia,
        "anticipo": anticipo, "saldo": saldo,
        "items": items, "subtotal": subtotal, "iva": iva, "total": total,
        "resolucion_dian": resolucion_dian, "res_fecha": res_fecha,
        "res_desde": res_desde, "res_hasta": res_hasta,
        "fecha_factura": date.today().isoformat(),
    }


def _excel_desde_datos(d: dict) -> bytes:
    """Genera el Excel de cotización a partir del set de datos."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cotización"

    fila = 1
    ws.cell(fila, 1, "COTIZACIÓN DE COMPRA")
    fila += 1
    ws.cell(fila, 1, _rnd_syn("numero_cotizacion_proveedor").title()); ws.cell(fila, 2, d["num_cot"])
    fila += 1
    ws.cell(fila, 1, _rnd_syn("proveedor_nombre").title()); ws.cell(fila, 2, d["empresa"])
    fila += 1
    ws.cell(fila, 1, _rnd_syn("proveedor_nit").upper()); ws.cell(fila, 2, d["nit"])
    fila += 1
    ws.cell(fila, 1, _rnd_syn("proveedor_email").title()); ws.cell(fila, 2, d["email"])
    fila += 1
    ws.cell(fila, 1, _rnd_syn("fecha_vigencia").title()); ws.cell(fila, 2, d["vigencia"])
    fila += 1
    ws.cell(fila, 1, _rnd_syn("forma_pago").title()); ws.cell(fila, 2, d["forma_pago"])
    fila += 1
    ws.cell(fila, 1, _rnd_syn("plazo_entrega").title()); ws.cell(fila, 2, d["plazo"])
    fila += 1
    ws.cell(fila, 1, _rnd_syn("garantia").title()); ws.cell(fila, 2, d["garantia"])
    fila += 1
    ws.cell(fila, 1, _rnd_syn("anticipo").title()); ws.cell(fila, 2, d["anticipo"])
    fila += 1
    ws.cell(fila, 1, _rnd_syn("pago_saldo").title()); ws.cell(fila, 2, d["saldo"])
    fila += 2

    ws.cell(fila, 1, "No.")
    ws.cell(fila, 2, _rnd_syn("descripcion").title())
    ws.cell(fila, 3, _rnd_syn("referencia").title())
    ws.cell(fila, 4, _rnd_syn("cantidad").title())
    ws.cell(fila, 5, _rnd_syn("valor_unitario").title())
    ws.cell(fila, 6, _rnd_syn("valor_total").title())
    fila += 1

    for it in d["items"]:
        ws.cell(fila, 1, it["num"]); ws.cell(fila, 2, it["desc"]); ws.cell(fila, 3, it["ref"])
        ws.cell(fila, 4, it["cant"]); ws.cell(fila, 5, it["precio"]); ws.cell(fila, 6, it["total"])
        fila += 1

    fila += 1
    ws.cell(fila, 5, _rnd_syn("valor_antes_iva").title()); ws.cell(fila, 6, d["subtotal"])
    fila += 1
    ws.cell(fila, 5, _rnd_syn("valor_iva").upper()); ws.cell(fila, 6, d["iva"])
    fila += 1
    ws.cell(fila, 5, _rnd_syn("valor_total").title()); ws.cell(fila, 6, d["total"])

    for col, w in [("A", 5), ("B", 40), ("C", 16), ("D", 10), ("E", 22), ("F", 18)]:
        ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _factura_pdf_desde_datos(d: dict) -> bytes:
    """Genera una factura de venta colombiana en PDF usando WeasyPrint."""
    try:
        from weasyprint import HTML as WeasyprintHTML
    except ImportError:
        raise RuntimeError("weasyprint no está instalado.")

    def _fmt(v: float) -> str:
        return f"$ {v:,.0f}".replace(",", ".")

    filas_items = "".join(
        f"""<tr>
          <td style="text-align:center">{it['num']}</td>
          <td>{it['desc']}</td>
          <td style="text-align:center">{it['ref']}</td>
          <td style="text-align:center">{it['cant']}</td>
          <td style="text-align:right">{_fmt(it['precio'])}</td>
          <td style="text-align:center">19%</td>
          <td style="text-align:right">{_fmt(it['total'])}</td>
        </tr>"""
        for it in d["items"]
    )

    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8"/>
<style>
  @page {{ size: letter; margin: 15mm 12mm; }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: Arial, sans-serif; font-size: 9pt; color: #1a1a1a; }}
  .header {{ display: flex; justify-content: space-between; align-items: flex-start;
             border-bottom: 2px solid #1a3a5c; padding-bottom: 8px; margin-bottom: 10px; }}
  .empresa-info h1 {{ font-size: 13pt; color: #1a3a5c; font-weight: bold; }}
  .empresa-info p {{ font-size: 8pt; color: #555; margin-top: 2px; }}
  .factura-box {{ border: 2px solid #1a3a5c; border-radius: 4px; padding: 8px 12px; text-align: center; min-width: 160px; }}
  .factura-box .tipo {{ font-size: 10pt; font-weight: bold; color: #1a3a5c; letter-spacing: 1px; }}
  .factura-box .numero {{ font-size: 14pt; font-weight: bold; color: #c0392b; margin: 4px 0; }}
  .factura-box .fecha {{ font-size: 8pt; color: #555; }}
  .dian-res {{ font-size: 7.5pt; color: #777; text-align: right; margin-top: 4px; }}
  .partes {{ display: flex; gap: 12px; margin-bottom: 10px; }}
  .parte-box {{ flex: 1; border: 1px solid #ccc; border-radius: 4px; padding: 7px 10px; }}
  .parte-box h3 {{ font-size: 8pt; font-weight: bold; color: #1a3a5c; text-transform: uppercase;
                  letter-spacing: 0.5px; border-bottom: 1px solid #eee; padding-bottom: 3px; margin-bottom: 5px; }}
  .parte-box p {{ font-size: 8.5pt; margin-bottom: 2px; }}
  .parte-box .nombre {{ font-weight: bold; font-size: 9.5pt; }}
  .condiciones {{ display: flex; gap: 12px; margin-bottom: 10px; }}
  .cond-item {{ flex: 1; background: #f5f7fa; border-radius: 3px; padding: 5px 8px; }}
  .cond-item .label {{ font-size: 7pt; color: #888; text-transform: uppercase; }}
  .cond-item .value {{ font-size: 8.5pt; font-weight: bold; color: #222; }}
  table {{ width: 100%; border-collapse: collapse; margin-bottom: 8px; }}
  thead tr {{ background: #1a3a5c; color: white; }}
  thead th {{ padding: 5px 6px; font-size: 8pt; text-align: left; }}
  tbody tr:nth-child(even) {{ background: #f8f9fb; }}
  tbody td {{ padding: 4px 6px; font-size: 8pt; border-bottom: 1px solid #eee; }}
  .totales {{ float: right; width: 260px; }}
  .totales table {{ width: 100%; }}
  .totales td {{ padding: 3px 8px; font-size: 8.5pt; }}
  .totales td:last-child {{ text-align: right; font-weight: 500; }}
  .totales .total-row td {{ background: #1a3a5c; color: white; font-weight: bold; font-size: 9.5pt; }}
  .clearfix::after {{ content: ""; display: table; clear: both; }}
  .footer {{ margin-top: 18px; border-top: 1px solid #ccc; padding-top: 8px;
             font-size: 7.5pt; color: #888; text-align: center; }}
  .cot-ref {{ font-size: 8pt; color: #666; margin-bottom: 8px;
              background: #fffbe6; border: 1px solid #ffe58f; border-radius: 3px; padding: 5px 10px; }}
</style>
</head>
<body>

<!-- ENCABEZADO -->
<div class="header">
  <div class="empresa-info">
    <h1>{d['empresa']}</h1>
    <p><strong>NIT:</strong> {d['nit']} — Responsable de IVA</p>
    <p><strong>Dir.:</strong> {d['direccion']}, {d['ciudad']}</p>
    <p><strong>Email:</strong> {d['email']}</p>
  </div>
  <div>
    <div class="factura-box">
      <div class="tipo">FACTURA DE VENTA</div>
      <div class="numero">{d['num_fact']}</div>
      <div class="fecha">Fecha: {d['fecha_factura']}</div>
      <div class="fecha">Ciudad: {d['ciudad']}</div>
    </div>
    <div class="dian-res">Res. DIAN N° {d['resolucion_dian']} del {d['res_fecha']}<br/>
      Facturas {d['res_desde']:,} al {d['res_hasta']:,}
    </div>
  </div>
</div>

<!-- REFERENCIA COTIZACIÓN -->
<div class="cot-ref">
  📄 Esta factura corresponde a la cotización <strong>{d['num_cot']}</strong> &nbsp;|&nbsp;
  Vigencia cotización: {d['vigencia']}
</div>

<!-- VENDEDOR / ADQUIRENTE -->
<div class="partes">
  <div class="parte-box">
    <h3>Vendedor / Proveedor</h3>
    <p class="nombre">{d['empresa']}</p>
    <p><strong>NIT:</strong> {d['nit']}</p>
    <p><strong>Dir.:</strong> {d['direccion']}</p>
    <p><strong>Ciudad:</strong> {d['ciudad']}</p>
    <p><strong>Email:</strong> {d['email']}</p>
  </div>
  <div class="parte-box">
    <h3>Adquirente</h3>
    <p class="nombre">{_ADQUIRENTE['nombre']}</p>
    <p><strong>NIT:</strong> {_ADQUIRENTE['nit']}</p>
    <p><strong>Dir.:</strong> {_ADQUIRENTE['direccion']}</p>
    <p><strong>Ciudad:</strong> {_ADQUIRENTE['ciudad']}</p>
    <p><strong>Email:</strong> {_ADQUIRENTE['email']}</p>
    <p><strong>Régimen:</strong> {_ADQUIRENTE['regimen']}</p>
  </div>
</div>

<!-- CONDICIONES -->
<div class="condiciones">
  <div class="cond-item"><div class="label">Forma de pago</div><div class="value">{d['forma_pago']}</div></div>
  <div class="cond-item"><div class="label">Plazo de entrega</div><div class="value">{d['plazo']}</div></div>
  <div class="cond-item"><div class="label">Garantía</div><div class="value">{d['garantia']}</div></div>
  <div class="cond-item"><div class="label">Anticipo</div><div class="value">{d['anticipo']}</div></div>
</div>

<!-- TABLA DE ÍTEMS -->
<table>
  <thead>
    <tr>
      <th style="width:30px">No.</th>
      <th>Descripción</th>
      <th>Referencia</th>
      <th style="width:45px;text-align:center">Cant.</th>
      <th style="text-align:right">Vlr. Unit.</th>
      <th style="width:40px;text-align:center">IVA</th>
      <th style="text-align:right">Total</th>
    </tr>
  </thead>
  <tbody>
    {filas_items}
  </tbody>
</table>

<!-- TOTALES -->
<div class="clearfix">
  <div class="totales">
    <table>
      <tr><td>Subtotal (sin IVA)</td><td>{_fmt(d['subtotal'])}</td></tr>
      <tr><td>IVA 19%</td><td>{_fmt(d['iva'])}</td></tr>
      <tr class="total-row"><td><strong>TOTAL A PAGAR</strong></td><td><strong>{_fmt(d['total'])}</strong></td></tr>
    </table>
  </div>
</div>

<!-- PIE -->
<div class="footer">
  Documento generado electrónicamente. Resolución DIAN N° {d['resolucion_dian']} del {d['res_fecha']},
  habilitando facturas del {d['res_desde']:,} al {d['res_hasta']:,} — {d['empresa']} NIT {d['nit']}
</div>

</body>
</html>"""

    pdf_bytes = WeasyprintHTML(string=html).write_pdf()
    return pdf_bytes
