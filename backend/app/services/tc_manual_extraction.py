"""Extracción de texto de manuales de funciones (T&C → SIG análisis IA)."""
from __future__ import annotations

import io
import logging
import os
import subprocess
import tempfile
from datetime import date, datetime, time
from typing import Any, Optional

log = logging.getLogger(__name__)

MAX_TEXT = 50_000
MIN_USEFUL = 50
MANUALES_DIR = "/app/data/tc_manuales"

_OLE_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


def manual_disk_path(cargo_id: int, manual_url: str) -> Optional[str]:
    """Resuelve ruta en disco (tc_manuales) a partir de la URL pública (/tc-manuales/)."""
    if not manual_url:
        return None
    ext = manual_url.rsplit(".", 1)[-1].lower() if "." in manual_url else ""
    if ext:
        primary = os.path.join(MANUALES_DIR, f"{cargo_id}.{ext}")
        if os.path.isfile(primary):
            return primary
    legacy = os.path.join("/app/data", manual_url.lstrip("/").replace("/", os.sep))
    if os.path.isfile(legacy):
        return legacy
    return None


def sniff_excel_ext(content: bytes, declared: str) -> str:
    """Corrige extensión cuando MIME/nombre no coincide con el binario real."""
    declared = declared.lower().lstrip(".")
    if content[:2] == b"PK":
        return "xlsx"
    if content[: len(_OLE_MAGIC)] == _OLE_MAGIC:
        return "xls"
    return declared


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M") if value.time() != time.min else value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    text = str(value).strip()
    if text.startswith("=") and len(text) > 1:
        return text[1:].strip()
    return text


def _extraer_xlsx(content: bytes) -> str:
    import openpyxl

    best = ""
    for data_only in (True, False):
        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=data_only, read_only=False)
            lines: list[str] = []
            for ws in wb.worksheets:
                lines.append(f"## {ws.title}")
                max_row = ws.max_row or 0
                max_col = ws.max_column or 0
                if max_row == 0 or max_col == 0:
                    continue
                for row in ws.iter_rows(
                    min_row=1,
                    max_row=max_row,
                    max_col=max_col,
                    values_only=True,
                ):
                    parts = [_cell_str(c) for c in row]
                    if any(parts):
                        lines.append("\t".join(parts))
            wb.close()
            text = "\n".join(lines).strip()
            if len(text) > len(best):
                best = text
            if len(text) >= MIN_USEFUL:
                return text[:MAX_TEXT]
        except Exception as exc:
            log.warning("[tc_manual] openpyxl data_only=%s: %s", data_only, exc)
    return best[:MAX_TEXT]


def _extraer_xls(content: bytes) -> str:
    try:
        import xlrd
    except ImportError:
        log.warning("[tc_manual] xlrd no instalado — .xls sin texto extraído")
        return ""

    lines: list[str] = []
    try:
        wb = xlrd.open_workbook(file_contents=content)
        for sheet in wb.sheets():
            lines.append(f"## {sheet.name}")
            for row_idx in range(sheet.nrows):
                parts = [_cell_str(sheet.cell_value(row_idx, col_idx)) for col_idx in range(sheet.ncols)]
                if any(parts):
                    lines.append("\t".join(parts))
    except Exception as exc:
        log.warning("[tc_manual] xlrd falló: %s", exc)
        return ""
    return "\n".join(lines).strip()[:MAX_TEXT]


def _extraer_excel(content: bytes, ext: str) -> str:
    ext = sniff_excel_ext(content, ext)
    if ext == "xlsx":
        text = _extraer_xlsx(content)
        if len(text.strip()) >= MIN_USEFUL:
            return text
        # Archivo .xls renombrado como .xlsx
        if content[: len(_OLE_MAGIC)] == _OLE_MAGIC:
            log.info("[tc_manual] .xlsx declarado pero binario OLE — probando xlrd")
            return _extraer_xls(content)
        return text
    if ext == "xls":
        text = _extraer_xls(content)
        if len(text.strip()) >= MIN_USEFUL:
            return text
        if content[:2] == b"PK":
            log.info("[tc_manual] .xls declarado pero binario ZIP — probando openpyxl")
            return _extraer_xlsx(content)
        return text
    return ""


def extraer_texto_manual(content: bytes, ext: str) -> str:
    ext = ext.lower().lstrip(".")
    try:
        if ext == "pdf":
            import pdfplumber

            with pdfplumber.open(io.BytesIO(content)) as pdf:
                digital = "\n".join(p.extract_text() or "" for p in pdf.pages)
            from app.services.ocr_service import texto_con_ocr_fallback

            return texto_con_ocr_fallback(digital, content)[:MAX_TEXT]

        if ext == "docx":
            from docx import Document

            doc = Document(io.BytesIO(content))
            return "\n".join(p.text for p in doc.paragraphs if p.text.strip())[:MAX_TEXT]

        if ext == "doc":
            return _extraer_doc_antiword(content)[:MAX_TEXT]

        if ext in ("xlsx", "xls", "xlsm"):
            return _extraer_excel(content, ext)
    except Exception as exc:
        log.warning("[tc_manual] Error extrayendo .%s: %s", ext, exc)
    return ""


def _extraer_doc_antiword(content: bytes) -> str:
    try:
        with tempfile.NamedTemporaryFile(suffix=".doc", delete=False) as tmp:
            tmp.write(content)
            tmp_path = tmp.name
        try:
            proc = subprocess.run(
                ["antiword", "-m", "UTF-8.txt", tmp_path],
                capture_output=True,
                text=True,
                timeout=120,
                check=False,
            )
            if proc.returncode == 0 and proc.stdout.strip():
                return proc.stdout
            log.warning("[tc_manual] antiword exit=%s stderr=%s", proc.returncode, proc.stderr[:200])
        finally:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass
    except FileNotFoundError:
        log.warning("[tc_manual] antiword no instalado — .doc sin texto extraído")
    except Exception as exc:
        log.warning("[tc_manual] antiword falló: %s", exc)
    return ""


def extraer_desde_archivo(path: str) -> str:
    ext = path.rsplit(".", 1)[-1].lower()
    with open(path, "rb") as f:
        content = f.read()
    if ext in ("xlsx", "xls", "xlsm"):
        return _extraer_excel(content, ext)
    return extraer_texto_manual(content, ext)


def cargo_manual_flags(manual_url: str, manual_text: str) -> dict[str, bool | int]:
    texto = (manual_text or "").strip()
    return {
        "tiene_archivo": bool(manual_url),
        "tiene_manual": len(texto) >= MIN_USEFUL,
        "texto_chars": len(texto),
    }
