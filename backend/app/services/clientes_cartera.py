"""Lógica compartida — cartera de clientes corporativos (Operativo + T&C lectura)."""
from __future__ import annotations

import io
import json
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from fastapi import HTTPException, UploadFile
from pydantic import BaseModel, Field
from sqlalchemy import text
from sqlmodel import Session, col, select

from app.models.sede import Sede
from app.personal_database import (
    PtcCliente,
    PtcClienteAnalista,
    PtcClienteAsignacion,
    PtcPersona,
    PtcTicketRol,
    get_personal_engine,
)
from app.routers.personal import _email_corporativo_efectivo

_ACTIVO = "Activo"
_ROLES_TICKET = ("supervisor", "analista", "coordinador")
_CONFIG_SEDES_INACTIVAS = "oper_clientes_sedes_inactivas"
_PLANTILLA_CLIENTES = (
    Path(__file__).resolve().parent.parent / "assets" / "Plantilla_Cargue_Clientes.xlsx"
)


class AsignacionBody(BaseModel):
    sede_id: int
    persona_id: Optional[int] = None


class ClienteBody(BaseModel):
    client_no: str = Field(min_length=1, max_length=50)
    dume_no: str = Field(default="", max_length=50)
    nombre: str = Field(min_length=1, max_length=200)
    activo: bool = True
    asignaciones: list[AsignacionBody] = Field(default_factory=list)
    # Analistas responsables para gestión de tickets (Zymo Ally) — distinto de
    # `asignaciones` (Cartera de Clientes, 1 persona por sede). Puede haber
    # varios a la vez.
    analistas_tickets: list[int] = Field(default_factory=list)


class ClienteUpdateBody(BaseModel):
    dume_no: Optional[str] = None
    nombre: Optional[str] = None
    activo: Optional[bool] = None
    asignaciones: Optional[list[AsignacionBody]] = None
    analistas_tickets: Optional[list[int]] = None


class SedesConfigBody(BaseModel):
    sedes_inactivas: list[int] = Field(default_factory=list)


class RolTicketBody(BaseModel):
    rol: str
    persona_ids: list[int] = Field(default_factory=list)


def get_personal_db():
    with Session(get_personal_engine()) as session:
        yield session


def _client_field(row: dict, keys: list[str]) -> str:
    for k in keys:
        for rk, rv in row.items():
            if str(rk).strip().lower() == k.lower():
                return str(rv or "").strip()
    return ""


def _sedes_inactivas(db: Session) -> set[int]:
    row = db.execute(
        text("SELECT value FROM ptc_config WHERE key=:k"),
        {"k": _CONFIG_SEDES_INACTIVAS},
    ).first()
    if not row or not row[0]:
        return set()
    try:
        data = json.loads(row[0])
        if isinstance(data, list):
            return {int(x) for x in data}
    except (TypeError, ValueError, json.JSONDecodeError):
        pass
    return set()


def _set_sedes_inactivas(db: Session, sede_ids: list[int]) -> None:
    payload = json.dumps(sorted(set(sede_ids)))
    db.execute(
        text("INSERT OR REPLACE INTO ptc_config (key, value) VALUES (:k, :v)"),
        {"k": _CONFIG_SEDES_INACTIVAS, "v": payload},
    )


def listar_sedes_cartera(main_db: Session, personal_db: Session) -> list[dict]:
    inactivas = _sedes_inactivas(personal_db)
    sedes = main_db.exec(select(Sede).order_by(col(Sede.name))).all()
    return [
        {
            "id": s.id,
            "nombre": s.name,
            "activa_en_cartera": s.id not in inactivas,
        }
        for s in sedes
    ]


def sedes_activas_ids(main_db: Session, personal_db: Session) -> list[int]:
    inactivas = _sedes_inactivas(personal_db)
    sedes = main_db.exec(select(Sede)).all()
    return [s.id for s in sedes if s.id not in inactivas]


def _cliente_dict(
    c: PtcCliente,
    db: Session,
    main_db: Session,
    *,
    sedes_activas: list[int] | None = None,
) -> dict:
    asigs = db.exec(
        select(PtcClienteAsignacion).where(PtcClienteAsignacion.cliente_id == c.id)
    ).all()
    asignaciones: dict[str, dict] = {}
    for a in asigs:
        if sedes_activas is not None and a.sede_id not in sedes_activas:
            continue
        persona = db.get(PtcPersona, a.persona_id) if a.persona_id else None
        sede = main_db.get(Sede, a.sede_id) if a.sede_id else None
        asignaciones[str(a.sede_id)] = {
            "sede_id": a.sede_id,
            "sede_nombre": sede.name if sede else "",
            "persona_id": a.persona_id,
            "persona_nombre": persona.nombre if persona else "",
            "habilitada": True,
        }
    if sedes_activas is not None:
        for sid in sedes_activas:
            key = str(sid)
            if key not in asignaciones:
                sede = main_db.get(Sede, sid)
                asignaciones[key] = {
                    "sede_id": sid,
                    "sede_nombre": sede.name if sede else "",
                    "persona_id": None,
                    "persona_nombre": "",
                    "habilitada": False,
                }
    analistas_tickets = [
        _persona_min(persona, main_db)
        for a in db.exec(select(PtcClienteAnalista).where(PtcClienteAnalista.cliente_id == c.id)).all()
        if (persona := db.get(PtcPersona, a.persona_id))
    ]
    return {
        "id": c.id,
        "client_no": c.client_no,
        "dume_no": c.dume_no,
        "nombre": c.nombre,
        "activo": c.activo,
        "asignaciones": asignaciones,
        "analistas_tickets": analistas_tickets,
        "created_at": c.created_at.isoformat(),
        "updated_at": c.updated_at.isoformat(),
    }


def _sync_asignaciones(db: Session, cliente_id: int, items: list[AsignacionBody]) -> None:
    existing = db.exec(
        select(PtcClienteAsignacion).where(PtcClienteAsignacion.cliente_id == cliente_id)
    ).all()
    for row in existing:
        db.delete(row)
    for item in items:
        db.add(
            PtcClienteAsignacion(
                cliente_id=cliente_id,
                sede_id=item.sede_id,
                persona_id=item.persona_id,
            )
        )


def _sync_analistas_tickets(db: Session, cliente_id: int, persona_ids: list[int]) -> None:
    existing = db.exec(
        select(PtcClienteAnalista).where(PtcClienteAnalista.cliente_id == cliente_id)
    ).all()
    for row in existing:
        db.delete(row)
    for persona_id in dict.fromkeys(persona_ids):  # dedupe, conserva orden
        db.add(PtcClienteAnalista(cliente_id=cliente_id, persona_id=persona_id))


def _persona_min(p: PtcPersona, main_db: Session) -> dict:
    return {
        "id": p.id,
        "nombre": p.nombre,
        "email": _email_corporativo_efectivo(p, main_db),
    }


def resolver_jerarquia_tickets(cliente_id: int, db: Session, main_db: Session) -> dict:
    """Para un cliente: sus analistas responsables + coordinador(es) +
    supervisor(es), caminando `jefe_directo_id` (no el organigrama de cargos,
    que es ambiguo — ver PtcPersona.jefe_directo_id)."""
    asignaciones = db.exec(
        select(PtcClienteAnalista).where(PtcClienteAnalista.cliente_id == cliente_id)
    ).all()
    analistas: list[dict] = []
    coordinadores: dict[int, dict] = {}
    supervisores: dict[int, dict] = {}
    for a in asignaciones:
        persona = db.get(PtcPersona, a.persona_id)
        if not persona:
            continue
        analistas.append(_persona_min(persona, main_db))
        coordinador = db.get(PtcPersona, persona.jefe_directo_id) if persona.jefe_directo_id else None
        if not coordinador:
            continue
        coordinadores[coordinador.id] = _persona_min(coordinador, main_db)
        supervisor = db.get(PtcPersona, coordinador.jefe_directo_id) if coordinador.jefe_directo_id else None
        if supervisor:
            supervisores[supervisor.id] = _persona_min(supervisor, main_db)
    return {
        "analistas": analistas,
        "coordinadores": list(coordinadores.values()),
        "supervisores": list(supervisores.values()),
    }


def _stats(clientes: list[PtcCliente], db: Session) -> dict:
    total = len(clientes)
    asignados = 0
    for c in clientes:
        rows = db.exec(
            select(PtcClienteAsignacion).where(
                PtcClienteAsignacion.cliente_id == c.id,
                col(PtcClienteAsignacion.persona_id).is_not(None),
            )
        ).all()
        if rows:
            asignados += 1
    return {
        "total": total,
        "asignados": asignados,
        "pendientes": total - asignados,
    }


def listar_clientes_response(
    db: Session,
    main_db: Session,
    *,
    q: Optional[str] = None,
    skip: int = 0,
    limit: int = 200,
) -> dict:
    query = select(PtcCliente).where(PtcCliente.activo == True)  # noqa: E712
    if q:
        term = f"%{q.lower()}%"
        query = query.where(
            col(PtcCliente.nombre).ilike(term)
            | col(PtcCliente.client_no).ilike(term)
            | col(PtcCliente.dume_no).ilike(term)
        )
    all_rows = db.exec(query.order_by(col(PtcCliente.nombre))).all()
    stats = _stats(all_rows, db)
    page = all_rows[skip : skip + limit]
    activas = sedes_activas_ids(main_db, db)
    return {
        **stats,
        "sedes_activas": activas,
        "items": [_cliente_dict(c, db, main_db, sedes_activas=activas) for c in page],
    }


def listar_clientes_simple(db: Session) -> list[dict]:
    """Lista liviana (id+nombre) para selects — no requiere mod_oper_clientes,
    a diferencia de listar_clientes_response (que trae asignaciones/dume)."""
    clientes = db.exec(
        select(PtcCliente).where(PtcCliente.activo == True).order_by(col(PtcCliente.nombre))  # noqa: E712
    ).all()
    return [{"id": c.id, "nombre": c.nombre} for c in clientes]


def listar_personas_simple(db: Session, main_db: Session, rol: Optional[str] = None) -> list[dict]:
    """Lista liviana (id+nombre+correo) de personas del Directorio para los
    selects de Supervisor/Analista/Coordinador del formulario de tickets. Con
    `rol`, solo devuelve las personas curadas para ese rol (ver PtcTicketRol y
    la pantalla "Configuración de Tickets" — un admin elige, por área, quién
    puede aparecer ahí, solo entre personas con cargo asignado). No requiere
    mod_tc, mismo criterio que listar_clientes_simple."""
    if rol:
        ids = [r.persona_id for r in db.exec(select(PtcTicketRol).where(PtcTicketRol.rol == rol)).all()]
        if not ids:
            return []
        personas = db.exec(
            select(PtcPersona).where(col(PtcPersona.id).in_(ids)).order_by(col(PtcPersona.nombre))
        ).all()
        return [_persona_min(p, main_db) for p in personas]
    personas = db.exec(
        select(PtcPersona).where(PtcPersona.estado == _ACTIVO).order_by(col(PtcPersona.nombre))
    ).all()
    return [_persona_min(p, main_db) for p in personas]


def listar_personas_con_cargo(
    db: Session, main_db: Session,
    q: Optional[str] = None, area_id: Optional[int] = None, cargo_id: Optional[int] = None,
) -> list[dict]:
    """Personas activas CON cargo asignado (empleados reales, no solo un
    registro suelto en el Directorio) — candidatos para la pantalla de
    curación de roles de ticket. Combinables: buscar por nombre/documento
    (mismo criterio que Formatos digitales, `persona-por-documento`), filtrar
    por área (`PtcPersona.area_id`) y/o por cargo específico.

    Nota (2026-07-27): el filtro por área se había descartado antes por
    "no encontrar candidatos" — la causa real era que nginx no proxeaba
    `/operativo/personas/*` al backend (devolvía el HTML de la SPA, no JSON),
    no un problema de este filtro. Ya corregido en nginx.conf.
    """
    query = select(PtcPersona).where(PtcPersona.estado == _ACTIVO, col(PtcPersona.cargo_id).is_not(None))
    if area_id is not None:
        query = query.where(PtcPersona.area_id == area_id)
    if cargo_id is not None:
        query = query.where(PtcPersona.cargo_id == cargo_id)
    personas = db.exec(query.order_by(col(PtcPersona.nombre))).all()
    if q:
        term = q.strip().lower()
        # .documento puede ser NULL en filas viejas/importadas aunque el modelo
        # declare default="" -- sin el `or ""` cualquier persona con documento
        # nulo revienta el list comprehension entero (AttributeError sobre
        # NoneType), dejando la búsqueda completa en 0 resultados en silencio.
        personas = [
            p for p in personas
            if term in p.nombre.lower() or term in (p.documento or "").lower()
        ]
    return [_persona_min(p, main_db) for p in personas[:100]]


def listar_cargos_simple(db: Session, area_id: Optional[int] = None) -> list[dict]:
    """Lista liviana (id+nombre) de cargos del Directorio, opcionalmente por
    área — para el filtro de Cargo en la curación de roles de ticket."""
    from app.personal_database import PtcCargo

    query = select(PtcCargo)
    if area_id is not None:
        query = query.where(PtcCargo.area_id == area_id)
    cargos = db.exec(query.order_by(col(PtcCargo.nombre))).all()
    return [{"id": c.id, "nombre": c.nombre} for c in cargos]


def resolver_persona_por_plataforma(db: Session, main_db: Session, plataforma: str, rol: str) -> list[dict]:
    """Dado el nombre de una Plataforma (= Sede real — los valores de
    ZymoConfigList listType=platforms en Zymo Ally coinciden por nombre con
    Sede.name, verificado 2026-07-27), busca entre las personas curadas para
    `rol` (PtcTicketRol) cuál pertenece a esa sede (PtcPersona.sede_id).
    Autocompleta "¿Quién gestiona el ticket?" (el supervisor de esa
    plataforma — el usuario confirmó que hay uno solo) y sugiere Coordinador
    sin depender de elegir un Cliente."""
    sede = main_db.exec(select(Sede).where(Sede.name == plataforma)).first()
    if not sede:
        return []
    ids_rol = [r.persona_id for r in db.exec(select(PtcTicketRol).where(PtcTicketRol.rol == rol)).all()]
    if not ids_rol:
        return []
    personas = db.exec(
        select(PtcPersona).where(col(PtcPersona.id).in_(ids_rol), PtcPersona.sede_id == sede.id)
    ).all()
    return [_persona_min(p, main_db) for p in personas]


def listar_roles_ticket(db: Session, main_db: Session) -> dict:
    """Personas curadas actualmente, agrupadas por rol de ticket."""
    result: dict[str, list[dict]] = {rol: [] for rol in _ROLES_TICKET}
    ids_por_rol: dict[str, list[int]] = {rol: [] for rol in _ROLES_TICKET}
    for r in db.exec(select(PtcTicketRol)).all():
        if r.rol in ids_por_rol:
            ids_por_rol[r.rol].append(r.persona_id)
    for rol, ids in ids_por_rol.items():
        if not ids:
            continue
        personas = db.exec(select(PtcPersona).where(col(PtcPersona.id).in_(ids))).all()
        result[rol] = [_persona_min(p, main_db) for p in personas]
    return result


def guardar_rol_ticket(db: Session, rol: str, persona_ids: list[int]) -> None:
    if rol not in _ROLES_TICKET:
        raise HTTPException(status_code=400, detail=f"Rol inválido: {rol}")
    existing = db.exec(select(PtcTicketRol).where(PtcTicketRol.rol == rol)).all()
    for row in existing:
        db.delete(row)
    for persona_id in dict.fromkeys(persona_ids):  # dedupe, conserva orden
        db.add(PtcTicketRol(persona_id=persona_id, rol=rol))
    db.commit()


def listar_analistas(db: Session) -> list[dict]:
    from app.personal_database import PtcCargo

    personas = db.exec(select(PtcPersona).where(PtcPersona.estado == _ACTIVO)).all()
    result = []
    for p in personas:
        if not p.cargo_id:
            continue
        cargo = db.get(PtcCargo, p.cargo_id)
        if not cargo:
            continue
        cn = cargo.nombre.lower()
        if "analista" in cn and "operac" in cn:
            result.append({"id": p.id, "nombre": p.nombre, "sede_id": p.sede_id})
    return result


def crear_cliente(body: ClienteBody, db: Session, main_db: Session) -> dict:
    dup = db.exec(
        select(PtcCliente).where(PtcCliente.client_no == body.client_no.strip())
    ).first()
    if dup:
        raise HTTPException(status_code=409, detail="Ya existe un cliente con ese número.")
    c = PtcCliente(
        client_no=body.client_no.strip(),
        dume_no=body.dume_no.strip(),
        nombre=body.nombre.strip(),
        activo=body.activo,
    )
    db.add(c)
    db.flush()
    _sync_asignaciones(db, c.id, body.asignaciones)
    _sync_analistas_tickets(db, c.id, body.analistas_tickets)
    db.commit()
    db.refresh(c)
    activas = sedes_activas_ids(main_db, db)
    return _cliente_dict(c, db, main_db, sedes_activas=activas)


def actualizar_cliente(
    cliente_id: int, body: ClienteUpdateBody, db: Session, main_db: Session
) -> dict:
    c = db.get(PtcCliente, cliente_id)
    if not c:
        raise HTTPException(status_code=404, detail="Cliente no encontrado.")
    if body.dume_no is not None:
        c.dume_no = body.dume_no.strip()
    if body.nombre is not None:
        c.nombre = body.nombre.strip()
    if body.activo is not None:
        c.activo = body.activo
    if body.asignaciones is not None:
        _sync_asignaciones(db, c.id, body.asignaciones)
    if body.analistas_tickets is not None:
        _sync_analistas_tickets(db, c.id, body.analistas_tickets)
    c.updated_at = datetime.utcnow()
    db.add(c)
    db.commit()
    db.refresh(c)
    activas = sedes_activas_ids(main_db, db)
    return _cliente_dict(c, db, main_db, sedes_activas=activas)


def eliminar_cliente(cliente_id: int, db: Session) -> None:
    c = db.get(PtcCliente, cliente_id)
    if not c:
        raise HTTPException(status_code=404, detail="Cliente no encontrado.")
    c.activo = False
    c.updated_at = datetime.utcnow()
    db.add(c)
    db.commit()


async def importar_excel(file: UploadFile, db: Session) -> dict:
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=422, detail="Formato no soportado. Usa .xlsx")
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=422, detail="No se pudo leer el Excel.") from exc

    ws = wb.active
    headers = [str(c.value or "").strip() for c in ws[1]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    created = updated = skipped = 0

    for row in rows:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            skipped += 1
            continue
        row_dict = dict(zip(headers, row))
        client_no = _client_field(
            row_dict,
            ["No de Cliente", "No Cliente", "Cliente", "N° Cliente", "Numero de Cliente"],
        )
        dume_no = _client_field(
            row_dict,
            ["No de DUME", "No DUME", "DUME", "N° DUME", "Numero de DUME"],
        )
        nombre = _client_field(
            row_dict,
            [
                "Nombre de Cliente",
                "Nombre Cliente",
                "Cliente Nombre",
                "Nombre o razón social",
                "Nombre",
                "Razon social",
                "Nombre o razon social",
            ],
        )
        if not client_no or not nombre:
            skipped += 1
            continue
        existing = db.exec(
            select(PtcCliente).where(PtcCliente.client_no == client_no)
        ).first()
        if existing:
            existing.dume_no = dume_no
            existing.nombre = nombre
            existing.activo = True
            existing.updated_at = datetime.utcnow()
            db.add(existing)
            updated += 1
        else:
            db.add(
                PtcCliente(
                    client_no=client_no,
                    dume_no=dume_no,
                    nombre=nombre,
                    activo=True,
                )
            )
            created += 1
    db.commit()
    return {"created": created, "updated": updated, "skipped": skipped}


def guardar_sedes_config(body: SedesConfigBody, db: Session, main_db: Session) -> list[dict]:
    all_ids = {s.id for s in main_db.exec(select(Sede)).all()}
    invalid = [x for x in body.sedes_inactivas if x not in all_ids]
    if invalid:
        raise HTTPException(status_code=400, detail="Sede no válida.")
    _set_sedes_inactivas(db, body.sedes_inactivas)
    db.commit()
    return listar_sedes_cartera(main_db, db)


def plantilla_path() -> Path:
    return _PLANTILLA_CLIENTES
