import io
from datetime import date

from sqlmodel import Session

from app.models.work_task import WorkTask
from app.schemas.task_dashboard import TaskFilters
from app.services.task_dashboard_service import get_team_tasks

_EXCEL_HEADERS = [
    "ID",
    "Fecha",
    "Responsable",
    "Título",
    "Descripción Técnica",
    "Etiqueta",
    "Plataforma",
    "Estado",
    "Hora Inicio",
    "Hora Cierre",
    "Minutos Totales",
]


def _task_to_row(task: WorkTask) -> list:
    return [
        task.id,
        str(task.fecha),
        task.subido_por_nombre,
        task.titulo,
        task.descripcion_tecnica,
        task.etiqueta,
        task.plataforma,
        task.estado,
        task.hora_inicio.isoformat() if task.hora_inicio else "",
        task.hora_cierre.isoformat() if task.hora_cierre else "",
        task.tiempo_total_minutos if task.tiempo_total_minutos is not None else "",
    ]


def build_tasks_excel(db: Session, filters: TaskFilters, owner_id: int) -> bytes:
    """Generates Excel file with filtered tasks. Uses openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    tasks = get_team_tasks(db, filters, owner_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Tareas"  # type: ignore[union-attr]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    ws.append(_EXCEL_HEADERS)  # type: ignore[union-attr]
    for cell in ws[1]:  # type: ignore[index]
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for task in tasks:
        ws.append(_task_to_row(task))  # type: ignore[union-attr]

    for col in ws.columns:  # type: ignore[union-attr]
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)  # type: ignore[union-attr]

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_tasks_pdf(db: Session, filters: TaskFilters, owner_id: int) -> bytes:
    """Generates PDF file with filtered tasks. Uses weasyprint + jinja2."""
    from jinja2 import Environment, BaseLoader
    from weasyprint import HTML

    tasks = get_team_tasks(db, filters, owner_id)

    _HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<style>
  body { font-family: Arial, sans-serif; font-size: 11px; margin: 20px; color: #222; }
  h1 { font-size: 16px; margin-bottom: 4px; }
  p.subtitle { font-size: 10px; color: #666; margin-bottom: 12px; }
  table { width: 100%; border-collapse: collapse; }
  thead { background-color: #2E4057; color: white; }
  th, td { border: 1px solid #ccc; padding: 4px 6px; text-align: left; }
  tr:nth-child(even) { background-color: #f5f5f5; }
  .estado-completada { color: #27ae60; font-weight: bold; }
  .estado-en_progreso { color: #2980b9; font-weight: bold; }
  .estado-bloqueada { color: #e74c3c; font-weight: bold; }
</style>
</head>
<body>
  <h1>Reporte de Tareas</h1>
  <p class="subtitle">Generado: {{ today }} | Total: {{ tasks|length }} tareas</p>
  <table>
    <thead>
      <tr>
        <th>Fecha</th>
        <th>Responsable</th>
        <th>Título</th>
        <th>Etiqueta</th>
        <th>Plataforma</th>
        <th>Estado</th>
        <th>Minutos</th>
      </tr>
    </thead>
    <tbody>
      {% for t in tasks %}
      <tr>
        <td>{{ t.fecha }}</td>
        <td>{{ t.subido_por_nombre }}</td>
        <td>{{ t.titulo }}</td>
        <td>{{ t.etiqueta }}</td>
        <td>{{ t.plataforma }}</td>
        <td class="estado-{{ t.estado }}">{{ t.estado }}</td>
        <td>{{ t.tiempo_total_minutos if t.tiempo_total_minutos is not none else '' }}</td>
      </tr>
      {% endfor %}
    </tbody>
  </table>
</body>
</html>
"""

    env = Environment(loader=BaseLoader())
    template = env.from_string(_HTML_TEMPLATE)
    html_str = template.render(tasks=tasks, today=str(date.today()))

    pdf_bytes = HTML(string=html_str).write_pdf()
    return pdf_bytes
