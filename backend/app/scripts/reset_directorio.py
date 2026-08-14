"""
Reset del directorio T&C (PtcPersona) contra el Excel de RRHH mas reciente.

El directorio en produccion tenia datos de la fase de pruebas, desactualizados.
Este script reemplaza esa informacion con "Bases de datos.xlsx" (hoja "Data
personal", 152 personas), cruzando el numero de carne contra la hoja
"Indefinidos" por documento.

Es un UPSERT por documento, NO un borrado y reinsercion:
  - Documento ya existe en ptc_persona -> se actualiza esa fila (mismo id).
    Esto es deliberado: otras tablas (capacitaciones, evaluaciones, sanciones,
    eventos de agenda, ptc_cliente_analista / ptc_cliente_asignacion que usa
    ZymoAlly para resolver jerarquia de tickets) referencian persona_id. Si se
    borrara y reinsertara con id nuevo, esas relaciones quedarian huerfanas
    en silencio.
  - Documento nuevo en el Excel -> se crea la persona.
  - Persona activa hoy cuyo documento NO aparece en el Excel nuevo -> se marca
    estado="Inactivo" (no se borra), preservando su historial relacionado.

La hoja "Temporales" (chip/SIM) NO se carga -- no existe campo para eso en
PtcPersona hoy. Se reporta solo como conteo informativo en el dry-run.

Uso (dentro del contenedor backend):
    python -m app.scripts.reset_directorio --file /app/data/import/base_datos.xlsx
    python -m app.scripts.reset_directorio --file /app/data/import/base_datos.xlsx --apply

Sin --apply corre en modo dry-run: no escribe nada, solo imprime el reporte.
"""
from __future__ import annotations

import argparse
import sys
from collections import Counter
from dataclasses import dataclass, field
from datetime import date
from typing import Optional

import openpyxl
from sqlmodel import Session, select

from app.database import get_engine
from app.models.area import Area as GlobalArea
from app.models.sede import Sede
from app.personal_database import (
    PtcCargo,
    PtcCargoSede,
    PtcPersona,
    create_personal_tables,
    get_personal_engine,
)


# ── Lectura del Excel ────────────────────────────────────────────────────────

@dataclass
class FilaPersona:
    documento: str
    nombre: str
    plataforma: str
    area: str
    cargo: str
    rh: str
    genero: str
    fecha_ingreso: Optional[date]
    fecha_nacimiento: Optional[date]
    correo: str
    celular: str
    tarjeta: str = ""
    tarjeta_fecha: Optional[date] = None
    es_temporal: bool = False


def _norm_documento(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return str(v).strip()


def _to_date(v) -> Optional[date]:
    if v is None or v == "":
        return None
    if hasattr(v, "date"):
        return v.date()
    return None


def leer_excel(path: str) -> tuple[list[FilaPersona], dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    dp = wb["Data personal"]
    temp = wb["Temporales"]
    indef = wb["Indefinidos"]

    # Indefinidos: documento -> (numero de carne, fecha de asignacion)
    carnet_por_doc: dict[str, tuple[str, Optional[date]]] = {}
    for r in range(2, indef.max_row + 1):
        doc = _norm_documento(indef.cell(row=r, column=1).value)
        if not doc:
            continue
        carnet = indef.cell(row=r, column=4).value
        fecha = _to_date(indef.cell(row=r, column=2).value)
        if carnet:
            carnet_por_doc[doc] = (str(carnet).strip(), fecha)

    # Temporales: solo para marcar tipo de contrato e informar conteo
    temporales_docs: set[str] = set()
    for r in range(2, temp.max_row + 1):
        doc = _norm_documento(temp.cell(row=r, column=1).value)
        if doc:
            temporales_docs.add(doc)

    filas: list[FilaPersona] = []
    for r in range(2, dp.max_row + 1):
        row = [dp.cell(row=r, column=c).value for c in range(1, 15)]
        if all(v is None for v in row):
            continue
        documento = _norm_documento(row[2])
        if not documento:
            continue
        carnet, carnet_fecha = carnet_por_doc.get(documento, ("", None))
        filas.append(FilaPersona(
            documento=documento,
            nombre=str(row[4] or "").strip(),
            plataforma=str(row[5] or "").strip(),
            area=str(row[6] or "").strip(),
            cargo=str(row[7] or "").strip(),
            rh=str(row[8] or "").strip(),
            genero=str(row[9] or "").strip().upper(),
            fecha_ingreso=_to_date(row[10]),
            fecha_nacimiento=_to_date(row[11]),
            correo=str(row[12] or "").strip(),
            celular=_norm_documento(row[13]),
            tarjeta=carnet,
            tarjeta_fecha=carnet_fecha,
            es_temporal=documento in temporales_docs,
        ))

    meta = {
        "total_data_personal": len(filas),
        "total_temporales": len(temporales_docs),
        "total_indefinidos": len(carnet_por_doc),
        "con_carnet": sum(1 for f in filas if f.tarjeta),
    }
    return filas, meta


# ── Matching contra catalogos existentes ─────────────────────────────────────

@dataclass
class Catalogos:
    sedes_por_nombre: dict[str, int] = field(default_factory=dict)
    areas_por_nombre: dict[str, int] = field(default_factory=dict)
    cargos_por_nombre: dict[str, PtcCargo] = field(default_factory=dict)


def cargar_catalogos(main_db: Session, personal_db: Session) -> Catalogos:
    cat = Catalogos()
    for s in main_db.exec(select(Sede)).all():
        cat.sedes_por_nombre[s.name.strip().upper()] = s.id
    for a in main_db.exec(select(GlobalArea)).all():
        cat.areas_por_nombre[a.name.strip().upper()] = a.id
    for c in personal_db.exec(select(PtcCargo)).all():
        cat.cargos_por_nombre[c.nombre.strip().upper()] = c
    return cat


def _initials(nombre: str) -> str:
    return "".join(w[0].upper() for w in nombre.split()[:2] if w)


def _tipo_contrato(fila: FilaPersona) -> str:
    if fila.es_temporal:
        return "Término fijo"
    if "APRENDIZ" in fila.cargo.upper():
        return "Aprendizaje SENA"
    return "Término indefinido"


# ── Aplicacion ────────────────────────────────────────────────────────────────

def procesar(
    filas: list[FilaPersona],
    cat: Catalogos,
    personal_db: Session,
    apply: bool,
) -> dict:
    plataformas_sin_match = Counter()
    areas_sin_match = Counter()
    cargos_pendientes: set[str] = set()  # normalizados, evita contar duplicados
    creados = 0
    actualizados = 0

    existentes = {
        p.documento: p
        for p in personal_db.exec(select(PtcPersona)).all()
        if p.documento
    }
    docs_nuevos = {f.documento for f in filas}

    for fila in filas:
        sede_id = cat.sedes_por_nombre.get(fila.plataforma.upper())
        if fila.plataforma and sede_id is None:
            plataformas_sin_match[fila.plataforma] += 1

        area_id = cat.areas_por_nombre.get(fila.area.upper())
        if fila.area and area_id is None:
            areas_sin_match[fila.area] += 1

        real_cargo_id = None
        if fila.cargo:
            cargo = cat.cargos_por_nombre.get(fila.cargo.upper())
            if not cargo:
                cargos_pendientes.add(fila.cargo.upper())
                if apply:
                    # Con area_id enlazado — el import/json viejo hacia este
                    # mismo paso pero sin enlazar area_id, quedaba vacio.
                    cargo = PtcCargo(nombre=fila.cargo, area_id=area_id)
                    personal_db.add(cargo)
                    personal_db.flush()
                    cat.cargos_por_nombre[fila.cargo.upper()] = cargo
            if cargo:
                real_cargo_id = cargo.id
                if apply and sede_id is not None:
                    link = personal_db.exec(
                        select(PtcCargoSede).where(
                            PtcCargoSede.cargo_id == real_cargo_id,
                            PtcCargoSede.sede_id == sede_id,
                        )
                    ).first()
                    if not link:
                        personal_db.add(PtcCargoSede(cargo_id=real_cargo_id, sede_id=sede_id))

        persona = existentes.get(fila.documento)
        if persona:
            actualizados += 1
        else:
            creados += 1

        if not apply:
            continue

        if persona is None:
            persona = PtcPersona(documento=fila.documento)
            existentes[fila.documento] = persona

        persona.nombre = fila.nombre
        persona.initials = _initials(fila.nombre)
        if sede_id is not None:
            persona.sede_id = sede_id
        if area_id is not None:
            persona.area_id = area_id
        if real_cargo_id is not None:
            persona.cargo_id = real_cargo_id
        persona.genero = fila.genero.capitalize() if fila.genero else persona.genero or ""
        persona.rh = fila.rh
        persona.tarjeta = fila.tarjeta or persona.tarjeta
        persona.tarjeta_fecha_asignacion = fila.tarjeta_fecha or persona.tarjeta_fecha_asignacion
        persona.email = fila.correo
        persona.telefono = fila.celular
        persona.fecha_nacimiento = fila.fecha_nacimiento
        persona.fecha_ingreso = fila.fecha_ingreso
        persona.tipo_contrato = _tipo_contrato(fila)
        persona.estado = "Activo"
        personal_db.add(persona)

    a_inactivar = [
        p for doc, p in existentes.items()
        if doc not in docs_nuevos and p.estado == "Activo"
    ]
    if apply:
        for p in a_inactivar:
            p.estado = "Inactivo"
            personal_db.add(p)
        personal_db.commit()

    return {
        "creados": creados,
        "actualizados": actualizados,
        "a_inactivar": len(a_inactivar),
        "nombres_a_inactivar": [p.nombre for p in a_inactivar],
        "plataformas_sin_match": plataformas_sin_match,
        "areas_sin_match": areas_sin_match,
        "cargos_nuevos": len(cargos_pendientes),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--file", required=True, help="Ruta al Excel 'Bases de datos.xlsx'")
    parser.add_argument("--apply", action="store_true", help="Escribe de verdad. Sin esto, solo dry-run.")
    args = parser.parse_args()

    filas, meta = leer_excel(args.file)

    print("=== Lectura del Excel ===")
    print(f"Data personal:            {meta['total_data_personal']} personas")
    print(f"Temporales (chip/SIM):    {meta['total_temporales']} filas -- NO se cargan (no hay campo en el directorio)")
    print(f"Indefinidos (carne):      {meta['total_indefinidos']} filas cruzadas por documento")
    print(f"Personas con carne:       {meta['con_carnet']} de {meta['total_data_personal']}")
    print()

    create_personal_tables()
    with Session(get_engine()) as main_db, Session(get_personal_engine()) as personal_db:
        cat = cargar_catalogos(main_db, personal_db)
        resultado = procesar(filas, cat, personal_db, apply=args.apply)

    print("=== " + ("APLICADO" if args.apply else "DRY-RUN (nada se escribio)") + " ===")
    print(f"Personas a crear:         {resultado['creados']}")
    print(f"Personas a actualizar:    {resultado['actualizados']}")
    print(f"Cargos nuevos a crear:    {resultado['cargos_nuevos']}")
    print(f"Personas a marcar Inactivo (no aparecen en el archivo nuevo): {resultado['a_inactivar']}")
    if resultado["nombres_a_inactivar"]:
        for n in resultado["nombres_a_inactivar"][:20]:
            print(f"    - {n}")
        if len(resultado["nombres_a_inactivar"]) > 20:
            print(f"    ... y {len(resultado['nombres_a_inactivar']) - 20} mas")
    print()

    if resultado["plataformas_sin_match"]:
        print("PLATAFORMA sin Sede correspondiente (revisar nombres en Configuracion > Plataformas):")
        for nombre, n in resultado["plataformas_sin_match"].most_common():
            print(f"    - '{nombre}': {n} personas")
        print()

    if resultado["areas_sin_match"]:
        print("AREA sin Area correspondiente (quedan sin area_id, revisar catalogo de Areas):")
        for nombre, n in resultado["areas_sin_match"].most_common():
            print(f"    - '{nombre}': {n} personas")
        print()

    if not args.apply:
        print("Nada se escribio. Corre de nuevo con --apply para aplicar los cambios.")
        if resultado["plataformas_sin_match"] or resultado["areas_sin_match"]:
            print("Recomendado: resolver los nombres sin match antes de --apply.")
            sys.exit(1)


if __name__ == "__main__":
    main()
